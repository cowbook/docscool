import os
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from flask import current_app

from .contracts_core import (
    CONTRACT_FIELD_KEYS,
    EXCEL_ALLOWED_EXTENSIONS,
    STAMP_TAX_RATE_BY_CONTRACT_TYPE,
    _department_dir,
    _format_decimal_plain,
    _get_department_names,
    _get_project_names,
    _match_option_value,
    _normalize_contract_file_path,
    _normalize_contract_type_value,
    _normalize_date_value,
    _parse_date,
    _safe_decimal,
)
from .models import Contract

_IMPORT_ERROR_REPORTS = {}

AMOUNT_UNIT_TO_WAN = {
    '元': Decimal('0.0001'),
    '千元': Decimal('0.1'),
    '万元': Decimal('1'),
    '万': Decimal('1'),
    '亿元': Decimal('10000'),
    '亿': Decimal('10000'),
}

EXCEL_HEADER_ALIASES = {
    'contract_name': ['合同名称', '合同名', '名称', '标题', '流程标题', '审批标题', '表单标题'],
    'contract_number': ['合同编号', '编号', '协议编号', '单号', '流程编号', '表单编号'],
    'contract_unit': ['合同单位', '对方单位', '签约单位', '相对方', '供应商', '供应商名称', '客户名称'],
    'contract_amount': ['合同金额', '金额', '价税合计', '含税金额', '总价', '合同总价', '合同总金额'],
    'handler': ['承办人', '经办人', '负责人', '申请人', '发起人', '填报人'],
    'handling_department': ['承办部门', '部门', '归口部门', '申请部门', '发起部门', '所属部门'],
    'contract_determination_method': ['合同确定方式', '确定方式', '采购方式', '招采方式', '定标方式'],
    'handling_date': ['承办日期', '处理日期', '审批日期', '日期', '发起日期', '申请日期', '创建时间', '提交时间'],
    'contract_type': ['合同类型', '类型'],
    'purchase_type': ['采购类型', '采购类别'],
    'stamp_tax_rate': ['印花税率', '税率'],
    'pricing_method': ['计价方式', '定价方式'],
    'copy_count': ['份数', '份数copy_count', '合同份数'],
    'save_place': ['存档位置', '归档位置', '保存位置', 'save_place'],
    'project': ['项目', '项目名称'],
}


EXCEL_HEADER_LOOKUP = {
    re.sub(r'[^\w\u4e00-\u9fff]+', '', alias.strip().lower()): field
    for field, aliases in EXCEL_HEADER_ALIASES.items()
    for alias in aliases
}

def _convert_amount_to_wan(number_text: str, unit_text: str) -> str:
    cleaned_number = (number_text or '').replace(',', '').replace('，', '').strip()
    amount = Decimal(cleaned_number)
    multiplier = AMOUNT_UNIT_TO_WAN.get((unit_text or '').strip(), Decimal('1'))
    return _format_decimal_plain(amount * multiplier)


def _normalize_excel_header(value) -> str:
    text = str(value or '').strip().lower()
    return re.sub(r'[^\w\u4e00-\u9fff]+', '', text)


def _stringify_excel_value(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return _format_decimal_plain(value)
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _format_decimal_plain(Decimal(str(value)))
    return str(value).strip()


def _is_excel_row_empty(row) -> bool:
    return not any(_stringify_excel_value(item).strip() for item in (row or []))


def _match_excel_field(header_text: str) -> str:
    normalized = _normalize_excel_header(header_text)
    if not normalized:
        return ''

    direct = EXCEL_HEADER_LOOKUP.get(normalized)
    if direct:
        return direct

    contains_rules = [
        ('contract_name', ('合同名称', '合同名', '流程标题', '审批标题', '表单标题', '标题')),
        ('contract_number', ('合同编号', '协议编号', '流程编号', '表单编号', '编号', '单号')),
        ('contract_unit', ('合同单位', '签订','对方单位', '签约单位', '相对方', '供应商', '客户名称')),
        ('contract_amount', ('合同金额', '合同总价', '合同总金额', '价税合计', '含税金额', '总价', '金额')),
        ('handler', ('承办人', '经办人', '负责人', '申请人', '发起人', '填报人')),
        ('handling_department', ('承办部门', '归口部门', '申请部门', '发起部门', '所属部门', '部门')),
        ('contract_determination_method', ('合同确定方式', '确定方式', '采购方式', '招采方式', '定标方式')),
        ('handling_date', ('承办日期', '日期','处理日期', '审批日期', '发起日期', '申请日期', '创建时间', '提交时间', '日期')),
        ('contract_type', ('合同类型',)),
        ('purchase_type', ('采购类型', '采购类别')),
        ('stamp_tax_rate', ('印花税率', '税率')),
        ('pricing_method', ('计价方式', '定价方式')),
        ('copy_count', ('份数', '份数copy_count', '合同份数')),
        ('save_place', ('存档位置', '归档位置', '保存位置', 'save_place')),
        ('is_archived', ('是否归档', '归档状态')),
        ('project', ('项目名称', '项目')),
    ]
    for field, keywords in contains_rules:
        if any(keyword == header_text for keyword in keywords):
            return field

    return ''


def _map_excel_columns(header_row):
    field_indexes = {}
    header_labels = {}

    for index, cell in enumerate(header_row or []):
        header_text = _stringify_excel_value(cell)
        field = _match_excel_field(header_text)
        if field and field not in field_indexes:
            field_indexes[field] = index
            header_labels[field] = header_text

    return field_indexes, header_labels


def _detect_excel_header(rows):
    best = None
    best_count = 0
    max_candidates = min(len(rows), 20)

    for index in range(max_candidates):
        row = rows[index]
        if _is_excel_row_empty(row):
            continue
        field_indexes, header_labels = _map_excel_columns(row)
        recognized_count = len(field_indexes)
        if recognized_count > best_count:
            best = (index, field_indexes, header_labels)
            best_count = recognized_count

    if best_count == 0:
        return None, {}, {}
    return best


def _detect_amount_unit_from_header(header_text: str) -> str:
    text = str(header_text or '')
    for unit in ('亿元', '万元', '千元', '万', '元'):
        if unit in text:
            return unit
    return '万元'


def _normalize_excel_amount(value, header_text: str) -> str:
    text = _stringify_excel_value(value)
    if not text:
        return ''

    normalized = text.replace(',', '').replace('，', '').strip()
    match = re.search(r'([+-]?[0-9]+(?:\.[0-9]+)?)(亿元|万元|千元|万|元)', normalized)
    if match:
        return _convert_amount_to_wan(match.group(1), match.group(2))

    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', normalized):
        return _convert_amount_to_wan(normalized, _detect_amount_unit_from_header(header_text))

    cleaned = re.sub(r'[^0-9.+-]', '', normalized)
    if cleaned.count('.') > 1:
        first_dot = cleaned.find('.')
        cleaned = cleaned[:first_dot + 1] + cleaned[first_dot + 1:].replace('.', '')
    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', cleaned):
        return _convert_amount_to_wan(cleaned, _detect_amount_unit_from_header(header_text))

    return ''


def _normalize_excel_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalize_date_value(_stringify_excel_value(value))


def _load_excel_sheets(uploaded_file):
    filename = (uploaded_file.filename or '').strip()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in EXCEL_ALLOWED_EXTENSIONS:
        raise ValueError('仅支持上传 xls 或 xlsx 文件')

    try:
        uploaded_file.stream.seek(0)
    except Exception:
        pass
    content = uploaded_file.stream.read()
    if not content:
        raise ValueError('Excel 文件内容为空')

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError('缺少 openpyxl 依赖，无法导入 xlsx 文件') from exc

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheets = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            sheets.append((sheet.title, rows))
        return sheets

    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError('缺少 xlrd 依赖，无法导入 xls 文件') from exc

    workbook = xlrd.open_workbook(file_contents=content)
    sheets = []
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            values = []
            for col_index in range(sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    cell_value = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
                    values.append(cell_value)
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    values.append(int(cell.value) if float(cell.value).is_integer() else cell.value)
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    values.append(bool(cell.value))
                elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                    values.append('')
                else:
                    values.append(cell.value)
            rows.append(values)
        sheets.append((sheet.name, rows))
    return sheets


def _build_contract_import_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '合同导入模板'

    headers = [
        '合同名称',
        '合同编号',
        '合同单位',
        '合同金额',
        '份数',
        '存档位置',
        '承办人',
        '承办部门',
        '合同确定方式',
        '承办日期',
        '合同类型',
        '采购类型',
        '印花税率',
        '计价方式',
        '项目',
    ]
    widths = [26, 20, 24, 16, 10, 20, 14, 18, 18, 14, 14, 14, 12, 14, 22]

    sheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='DCE6F1')
    header_font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = widths[index - 1]

    sheet.freeze_panes = 'A2'

    notes = workbook.create_sheet('填写说明')
    note_rows = [
        ['说明', '内容'],
        ['必要列', '合同名称、承办部门'],
        ['金额规则', '请填写元单位的纯数字，可保留 8 位及以上小数'],
        ['份数规则', '选填，纯数字（整数）'],
        ['存档位置规则', '选填，最多50个字符'],
        ['日期格式', '建议使用 YYYY-MM-DD，例如 2026-04-03'],
        ['承办部门', '必须填写系统中已经配置的部门名称'],
        ['项目', '如填写，必须填写系统中已配置的项目名称'],
        ['支持格式', 'xls、xlsx'],
    ]
    for row in note_rows:
        notes.append(row)
    notes.column_dimensions['A'].width = 18
    notes.column_dimensions['B'].width = 80

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _build_import_error_report(sheet_name: str, source_headers, failed_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '导入失败明细'

    normalized_headers = [str(item or '').strip() for item in (source_headers or [])]
    if not normalized_headers:
        normalized_headers = [f'原始列{i + 1}' for i in range(max((len(item.get('row_values') or []) for item in failed_rows), default=0))]

    headers = ['Excel工作表', 'Excel行号'] + normalized_headers + ['失败原因']
    widths = [18, 12] + [max(12, min(28, len(header) + 4)) for header in normalized_headers] + [40]

    sheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
    header_font = Font(bold=True)
    for index, _header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = widths[index - 1]

    for item in failed_rows:
        row_values = [_stringify_excel_value(value) for value in (item.get('row_values') or [])]
        padded_values = row_values + [''] * max(0, len(normalized_headers) - len(row_values))
        sheet.append([
            item.get('sheet_name') or sheet_name,
            item.get('row', ''),
            *padded_values[:len(normalized_headers)],
            item.get('message', ''),
        ])

    sheet.freeze_panes = 'A2'
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _store_import_error_report(sheet_name: str, source_headers, failed_rows):
    if not failed_rows:
        return '', ''

    output = _build_import_error_report(sheet_name, source_headers, failed_rows)
    token = uuid4().hex
    filename = f'合同导入失败明细_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    _IMPORT_ERROR_REPORTS[token] = {
        'filename': filename,
        'content': output.getvalue(),
    }
    return token, filename


def _normalize_excel_option_fields(fields: dict, option_sets: dict) -> dict:
    normalized = dict(fields or {})

    normalized['handling_department'] = _match_option_value(
        normalized.get('handling_department', ''),
        option_sets.get('handling_department', []),
        (normalized.get('handling_department') or '').strip(),
    )
    normalized['project'] = _match_option_value(
        normalized.get('project', ''),
        option_sets.get('project', []),
        (normalized.get('project') or '').strip(),
    )
    normalized['contract_determination_method'] = _match_option_value(
        normalized.get('contract_determination_method', ''),
        option_sets.get('contract_determination_method', []),
        (normalized.get('contract_determination_method') or '').strip(),
    )
    normalized['contract_type'] = _match_option_value(
        normalized.get('contract_type', ''),
        option_sets.get('contract_type', []),
        (normalized.get('contract_type') or '').strip(),
    )
    normalized['purchase_type'] = _match_option_value(
        normalized.get('purchase_type', ''),
        option_sets.get('purchase_type', []),
        (normalized.get('purchase_type') or '').strip(),
    )
    normalized['pricing_method'] = _match_option_value(
        normalized.get('pricing_method', ''),
        option_sets.get('pricing_method', []),
        (normalized.get('pricing_method') or '').strip(),
    )
    normalized['is_archived'] = _match_option_value(
        normalized.get('is_archived', ''),
        option_sets.get('is_archived', []),
        (normalized.get('is_archived') or '').strip(),
    )
    return normalized


def _build_import_payload_from_row(row, field_indexes, header_labels, option_sets):
    payload = {key: '' for key in CONTRACT_FIELD_KEYS}

    for field, index in field_indexes.items():
        raw_value = row[index] if index < len(row) else ''
        if field == 'contract_amount':
            payload[field] = _stringify_excel_value(raw_value)
        elif field == 'handling_date':
            payload[field] = _normalize_excel_date(raw_value)
        else:
            payload[field] = _stringify_excel_value(raw_value)

    payload['contract_name'] = payload['contract_name'].replace('\n', ' ').strip()
    payload['contract_number'] = payload['contract_number'].strip()
    payload['contract_unit'] = payload['contract_unit'].strip()
    payload['copy_count'] = payload.get('copy_count', '').strip()
    payload['save_place'] = payload.get('save_place', '').strip()

    return _normalize_excel_option_fields(payload, option_sets)

def _build_contract_record(body: dict, created_by: str, pending_contract_numbers=None, update_mode=True):
    payload = body or {}
    has_file_path_input = 'file_path' in payload or 'path' in payload
    raw_file_path = payload.get('file_path', payload.get('path'))
    normalized_file_path = _normalize_contract_file_path(raw_file_path) if has_file_path_input else None
    normalized_contract_type = _normalize_contract_type_value(payload.get('contract_type'))
    normalized_stamp_tax_rate = (payload.get('stamp_tax_rate') or '').strip() or STAMP_TAX_RATE_BY_CONTRACT_TYPE.get(normalized_contract_type, '')

    required = ['contract_name', 'handling_department']
    missing = [key for key in required if not str(payload.get(key, '')).strip()]
    if missing:
        return None, f'Missing required fields: {", ".join(missing)}', 400, False

    contract_amount_text = str(payload.get('contract_amount') or '').strip()
    amount = _safe_decimal(contract_amount_text) if contract_amount_text else None
    if contract_amount_text and amount is None:
        if update_mode:
            amount = None
        else:
            return None, 'contract_amount is invalid', 400, False

    copy_count_text = str(payload.get('copy_count') or '').strip()
    if copy_count_text and not re.fullmatch(r'\d+', copy_count_text):
        if update_mode:
            copy_count = None
        else:
            return None, 'copy_count is invalid', 400, False
    else:
        copy_count = int(copy_count_text) if copy_count_text else None

    save_place_text = str(payload.get('save_place') or '').strip()
    if len(save_place_text) > 50:
        return None, 'save_place is too long (max 50)', 400, False
    save_place = save_place_text or None

    contract_number = (payload.get('contract_number') or '').strip()
    if not contract_number:
        return None, '合同编号不能为空', 400, False
    
    pending_contract_numbers = pending_contract_numbers or set()
    is_update = False
    
    if contract_number:
        if contract_number in pending_contract_numbers and not update_mode:
            return None, 'contract_number already exists', 409, False
        
        existing_contract = Contract.query.filter_by(contract_number=contract_number).first()
        if existing_contract:
            # Allow update if contract_number exists but belongs to the same record (e.g. during import with multiple rows sharing the same contract_number)
            if update_mode:
                incoming_contract_name = (payload.get('contract_name') or '').strip()
                existing_contract_name = (existing_contract.contract_name or '').strip()
                if incoming_contract_name != existing_contract_name:
                    return None, 'contract_number already exists with different contract_name', 409, False

                '''
                if existing_contract.is_archived == '已归档':
                    return None, '已归档的合同只能由管理员进行修改', 403, False
                '''
                existing_contract.contract_name = incoming_contract_name
                existing_contract.contract_unit = (payload.get('contract_unit') or '').strip() or None
                existing_contract.amount = amount
                existing_contract.currency = 'CNY'
                existing_contract.handler = (payload.get('handler') or '').strip() or None
                if has_file_path_input:
                    existing_contract.file_path = normalized_file_path or None
                existing_contract.department = (payload.get('handling_department') or '').strip()
                existing_contract.contract_determination_method = (payload.get('contract_determination_method') or '').strip() or None
                existing_contract.handling_date = _parse_date(payload.get('handling_date'))
                existing_contract.contract_type = normalized_contract_type or None
                existing_contract.purchase_type = (payload.get('purchase_type') or '').strip() or None
                existing_contract.stamp_tax_rate = normalized_stamp_tax_rate or None
                existing_contract.pricing_method = (payload.get('pricing_method') or '').strip() or None
                existing_contract.copy_count = copy_count
                existing_contract.save_place = save_place
                existing_contract.project = (payload.get('project') or '').strip() or None
                existing_contract.fullbody = (payload.get('fullbody') or '').strip() or None
                existing_contract.start_date = _parse_date(payload.get('start_date'))
                existing_contract.end_date = _parse_date(payload.get('end_date'))
                existing_contract.status = (payload.get('status') or 'active').strip() or 'active'
                existing_contract.updated_by = created_by
                return existing_contract, '', 0, True
            else:
                return None, 'contract_number already exists', 409, False

    department = (payload.get('handling_department') or '').strip()
    if not update_mode:
        allowed_departments = _get_department_names()
        if department not in allowed_departments:
            return None, 'handling_department is not in configured department settings', 400, False
    _department_dir(department)

    project = (payload.get('project') or '').strip() or None
    if project:
        allowed_projects = _get_project_names()
        if project not in allowed_projects:
            return None, 'project is not in configured project settings', 400, False

    record = Contract(
        contract_number=contract_number or None,
        contract_name=(payload.get('contract_name') or '').strip(),
        contract_unit=(payload.get('contract_unit') or '').strip() or None,
        amount=amount,
        currency='CNY',
        handler=(payload.get('handler') or '').strip() or None,
        department=department,
        contract_determination_method=(payload.get('contract_determination_method') or '').strip() or None,
        handling_date=_parse_date(payload.get('handling_date')),
        contract_type=normalized_contract_type or None,
        purchase_type=(payload.get('purchase_type') or '').strip() or None,
        stamp_tax_rate=normalized_stamp_tax_rate or None,
        pricing_method=(payload.get('pricing_method') or '').strip() or None,
        copy_count=copy_count,
        save_place=save_place,
        is_archived='未归档',
        project=project,
        file_path=normalized_file_path or None,
        fullbody=(payload.get('fullbody') or '').strip() or None,
        start_date=_parse_date(payload.get('start_date')),
        end_date=_parse_date(payload.get('end_date')),
        status=(payload.get('status') or 'active').strip() or 'active',
        created_by=created_by,
        updated_by=created_by,
    )
    return record, '', 0, False
