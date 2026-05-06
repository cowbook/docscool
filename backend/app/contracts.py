# --- 文件末尾追加 ---
# 仪表盘统计接口，需在 contracts_bp 定义后声明
import os
import posixpath
import re
import json
import time
import getpass
import mimetypes

from difflib import SequenceMatcher
from io import BytesIO
from uuid import uuid4
from urllib.parse import unquote, urlparse
from email.utils import format_datetime
from datetime import timezone
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import requests
from sqlalchemy import String, cast, func, or_
from flask import Blueprint, current_app, g, jsonify, request, send_file
from .auth import get_cached_user_password, require_auth
from .extensions import db
from .models import Contract, Department, ProjectOption
from .ocr_utils import (
    extract_ai_content_from_pdf,
    extract_pdf_text,
    extract_pdf_text_via_ocr,
    get_ocr_engine,
    mineru_extract_text_from_uploaded_pdf,
    mineru_auth_headers,
    _preview_lines,
)

contracts_bp = Blueprint('contracts', __name__, url_prefix='/api')
_IMPORT_ERROR_REPORTS = {}
EXTERNAL_API_TIMEOUT_SECONDS = 300
DEFAULT_DEPARTMENT_NAME = '财务部'



CONTRACT_FIELD_KEYS = [
    'contract_name',
    'contract_number',
    'contract_unit',
    'contract_amount',
    'handler',
    'handling_department',
    'contract_determination_method',
    'handling_date',
    'contract_type',
    'purchase_type',
    'stamp_tax_rate',
    'pricing_method',
    'is_archived',
    'project',
]

STAMP_TAX_RATE_BY_CONTRACT_TYPE = {
    '买卖合同': '0.03%',
    '借款合同': '0.005%',
    '租赁合同': '0.1%',
    '承揽合同': '0.03%',
    '建设工程合同': '0.03%',
    '运输合同': '0.03%',
    '技术合同': '0.03%',
    '保管合同': '0.1%',
    '仓储合同': '0.1%',
    '财产保险合同': '0.1%',
    '人力资源': '',
    '其它': '',
}

LEGACY_INVALID_CONTRACT_TYPES = {
    '会议承办',
    '会议承办合同',
    '工程类',
    '采购类',
    '服务类',
}

OPTION_FIELD_DEFAULTS = {
    'project': '无',
    'contract_determination_method': '直接采购',
    'is_archived': '已归档',
}


CSV_OPTION_DEFAULTS = {
    'contract_determination_method': ['询比采购', '竞价采购', '谈判采购', '直接采购','非采购类'],
    'contract_type': ['买卖合同', '借款合同', '租赁合同', '承揽合同', '建设工程合同', '运输合同', '技术合同', '保管合同', '仓储合同', '财产保险合同','人力资源','其它'],
    'purchase_type': ['工程', '服务', '设备采购', '非采购类'],
    'pricing_method': ['单价合同', '总价合同','其他'],
    'is_archived': ['已归档', '未归档'],
}


SYNOLOGY_AUTH_ERROR_MESSAGES = {
    400: '账号或密码错误',
    401: '账号被停用',
    402: '账号权限不足',
    403: '需要二步验证(OTP)',
    404: '二步验证(OTP)错误',
    407: 'IP 被阻止，请稍后重试',
}


SYNOLOGY_COMMON_ERROR_MESSAGES = {
    100: '未知错误',
    101: '参数错误',
    102: 'API 不存在',
    103: '方法不存在',
    104: 'API 版本不支持',
    105: '权限不足',
    106: '会话已超时，请重新登录',
    107: '会话被中断，请重新登录',
    119: '会话无效或不存在，请检查上传会话配置并重新登录',
}


SYNOLOGY_FILESTATION_ERROR_MESSAGES = {
    400: '上传参数错误',
    401: '文件路径非法',
    402: '系统繁忙，请稍后再试',
    403: '无上传权限',
    404: '目录不存在',
    405: '文件已存在且不允许覆盖',
    406: '存储配额不足',
    407: '存储空间不足',
    408: '文件系统为只读',
    418: '文件名或路径包含非法字符',
    414: '目录已存在',
    415: '目录不存在',
}


AMOUNT_UNIT_TO_WAN = {
    '元': Decimal('0.0001'),
    '千元': Decimal('0.1'),
    '万元': Decimal('1'),
    '万': Decimal('1'),
    '亿元': Decimal('10000'),
    '亿': Decimal('10000'),
}

AI_AMOUNT_UNIT_TO_YUAN = {
    '元': Decimal('1'),
    '千元': Decimal('1000'),
    '万元': Decimal('10000'),
    '万': Decimal('10000'),
    '亿元': Decimal('100000000'),
    '亿': Decimal('100000000'),
}


EXCEL_ALLOWED_EXTENSIONS = {'.xls', '.xlsx'}


AI_MATCH_CANDIDATE_LIMIT = 20


EXCEL_HEADER_ALIASES = {
    'contract_name': ['合同名称', '合同名', '名称', '标题', '流程标题', '审批标题', '表单标题'],
    'contract_number': ['合同编号', '编号', '协议编号', '单号', '流程编号', '表单编号'],
    'contract_unit': ['合同单位', '对方单位', '签约单位', '相对方', '供应商', '供应商名称', '客户名称'],
    'contract_amount': ['合同金额', '金额', '价税合计', '含税金额', '总价', '合同总价', '合同总金额'],
    'handler': ['承办人', '经办人', '负责人', '申请人', '发起人', '填报人'],
    'handling_department': ['承办部门', '部门', '归口部门', '申请部门', '发起部门', '所属部门'],
    'contract_determination_method': ['合同确定方式', '确定方式', '采购方式', '招采方式', '定标方式'],
    'handling_date': ['承办日期', '处理日期', '审批日期', '日期', '发起日期', '申请日期', '创建时间', '提交时间'],
    'contract_type': ['合同类型', '类型'],
    'purchase_type': ['采购类型', '采购类别'],
    'stamp_tax_rate': ['印花税率', '税率'],
    'pricing_method': ['计价方式', '定价方式'],
    'copy_count': ['份数', '份数copy_count', '合同份数'],
    'save_place': ['存档位置', '归档位置', '保存位置', 'save_place'],
    'project': ['项目', '项目名称'],
}


EXCEL_HEADER_LOOKUP = {
    re.sub(r'[^\w\u4e00-\u9fff]+', '', alias.strip().lower()): field
    for field, aliases in EXCEL_HEADER_ALIASES.items()
    for alias in aliases
}


def _parse_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _safe_decimal(value):
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _normalize_contract_type_value(value: str) -> str:
    text = (value or '').strip()
    if not text:
        return ''
    if text in LEGACY_INVALID_CONTRACT_TYPES:
        return ''
    return text


def _format_decimal_plain(value: Decimal) -> str:
    text = format(value, 'f')
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    return text or '0'


def _convert_amount_to_wan(number_text: str, unit_text: str) -> str:
    cleaned_number = (number_text or '').replace(',', '').replace('，', '').strip()
    amount = Decimal(cleaned_number)
    multiplier = AMOUNT_UNIT_TO_WAN.get((unit_text or '').strip(), Decimal('1'))
    return _format_decimal_plain(amount * multiplier)


def _normalize_excel_header(value) -> str:
    text = str(value or '').strip().lower()
    return re.sub(r'[^\w\u4e00-\u9fff]+', '', text)


def _stringify_excel_value(value) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return _format_decimal_plain(value)
    if isinstance(value, bool):
        return '是' if value else '否'
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return _format_decimal_plain(Decimal(str(value)))
    return str(value).strip()


def _is_excel_row_empty(row) -> bool:
    return not any(_stringify_excel_value(item).strip() for item in (row or []))


def _match_excel_field(header_text: str) -> str:
    normalized = _normalize_excel_header(header_text)
    if not normalized:
        return ''

    direct = EXCEL_HEADER_LOOKUP.get(normalized)
    if direct:
        return direct

    contains_rules = [
        ('contract_name', ('合同名称', '合同名', '流程标题', '审批标题', '表单标题', '标题')),
        ('contract_number', ('合同编号', '协议编号', '流程编号', '表单编号', '编号', '单号')),
        ('contract_unit', ('合同单位', '签订','对方单位', '签约单位', '相对方', '供应商', '客户名称')),
        ('contract_amount', ('合同金额', '合同总价', '合同总金额', '价税合计', '含税金额', '总价', '金额')),
        ('handler', ('承办人', '经办人', '负责人', '申请人', '发起人', '填报人')),
        ('handling_department', ('承办部门', '归口部门', '申请部门', '发起部门', '所属部门', '部门')),
        ('contract_determination_method', ('合同确定方式', '确定方式', '采购方式', '招采方式', '定标方式')),
        ('handling_date', ('承办日期', '日期','处理日期', '审批日期', '发起日期', '申请日期', '创建时间', '提交时间', '日期')),
        ('contract_type', ('合同类型',)),
        ('purchase_type', ('采购类型', '采购类别')),
        ('stamp_tax_rate', ('印花税率', '税率')),
        ('pricing_method', ('计价方式', '定价方式')),
        ('copy_count', ('份数', '份数copy_count', '合同份数')),
        ('save_place', ('存档位置', '归档位置', '保存位置', 'save_place')),
        ('is_archived', ('是否归档', '归档状态')),
        ('project', ('项目名称', '项目')),
    ]
    for field, keywords in contains_rules:
        if any(keyword == header_text for keyword in keywords):
            return field

    return ''


def _map_excel_columns(header_row):
    field_indexes = {}
    header_labels = {}

    for index, cell in enumerate(header_row or []):
        header_text = _stringify_excel_value(cell)
        field = _match_excel_field(header_text)
        if field and field not in field_indexes:
            field_indexes[field] = index
            header_labels[field] = header_text

    return field_indexes, header_labels


def _detect_excel_header(rows):
    best = None
    best_count = 0
    max_candidates = min(len(rows), 20)

    for index in range(max_candidates):
        row = rows[index]
        if _is_excel_row_empty(row):
            continue
        field_indexes, header_labels = _map_excel_columns(row)
        recognized_count = len(field_indexes)
        if recognized_count > best_count:
            best = (index, field_indexes, header_labels)
            best_count = recognized_count

    if best_count == 0:
        return None, {}, {}
    return best


def _detect_amount_unit_from_header(header_text: str) -> str:
    text = str(header_text or '')
    for unit in ('亿元', '万元', '千元', '万', '元'):
        if unit in text:
            return unit
    return '万元'


def _normalize_excel_amount(value, header_text: str) -> str:
    text = _stringify_excel_value(value)
    if not text:
        return ''

    normalized = text.replace(',', '').replace('，', '').strip()
    match = re.search(r'([+-]?[0-9]+(?:\.[0-9]+)?)(亿元|万元|千元|万|元)', normalized)
    if match:
        return _convert_amount_to_wan(match.group(1), match.group(2))

    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', normalized):
        return _convert_amount_to_wan(normalized, _detect_amount_unit_from_header(header_text))

    cleaned = re.sub(r'[^0-9.+-]', '', normalized)
    if cleaned.count('.') > 1:
        first_dot = cleaned.find('.')
        cleaned = cleaned[:first_dot + 1] + cleaned[first_dot + 1:].replace('.', '')
    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', cleaned):
        return _convert_amount_to_wan(cleaned, _detect_amount_unit_from_header(header_text))

    return ''


def _normalize_excel_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalize_date_value(_stringify_excel_value(value))


def _load_excel_rows(uploaded_file):
    filename = (uploaded_file.filename or '').strip()
    ext = os.path.splitext(filename)[1].lower()
    if ext not in EXCEL_ALLOWED_EXTENSIONS:
        raise ValueError('仅支持上传 xls 或 xlsx 文件')

    try:
        uploaded_file.stream.seek(0)
    except Exception:
        pass
    content = uploaded_file.stream.read()
    if not content:
        raise ValueError('Excel 文件内容为空')

    if ext == '.xlsx':
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError('缺少 openpyxl 依赖，无法导入 xlsx 文件') from exc

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook.worksheets[0]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return sheet.title, rows

    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError('缺少 xlrd 依赖，无法导入 xls 文件') from exc

    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    rows = []
    for row_index in range(sheet.nrows):
        values = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                cell_value = xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
                values.append(cell_value)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                values.append(int(cell.value) if float(cell.value).is_integer() else cell.value)
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                values.append(bool(cell.value))
            elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                values.append('')
            else:
                values.append(cell.value)
        rows.append(values)
    return sheet.name, rows


def _build_contract_import_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '合同导入模板'

    headers = [
        '合同名称',
        '合同编号',
        '合同单位',
        '合同金额',
        '份数',
        '存档位置',
        '承办人',
        '承办部门',
        '合同确定方式',
        '承办日期',
        '合同类型',
        '采购类型',
        '印花税率',
        '计价方式',
        '项目',
    ]
    widths = [26, 20, 24, 16, 10, 20, 14, 18, 18, 14, 14, 14, 12, 14, 22]

    sheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='DCE6F1')
    header_font = Font(bold=True)
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = widths[index - 1]

    sheet.freeze_panes = 'A2'

    notes = workbook.create_sheet('填写说明')
    note_rows = [
        ['说明', '内容'],
        ['必要列', '合同名称、承办部门'],
        ['金额规则', '请填写元单位的纯数字，可保留 8 位及以上小数'],
        ['份数规则', '选填，纯数字（整数）'],
        ['存档位置规则', '选填，最多50个字符'],
        ['日期格式', '建议使用 YYYY-MM-DD，例如 2026-04-03'],
        ['承办部门', '必须填写系统中已经配置的部门名称'],
        ['项目', '如填写，必须填写系统中已配置的项目名称'],
        ['支持格式', 'xls、xlsx'],
    ]
    for row in note_rows:
        notes.append(row)
    notes.column_dimensions['A'].width = 18
    notes.column_dimensions['B'].width = 80

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _build_import_error_report(sheet_name: str, source_headers, failed_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '导入失败明细'

    normalized_headers = [str(item or '').strip() for item in (source_headers or [])]
    if not normalized_headers:
        normalized_headers = [f'原始列{i + 1}' for i in range(max((len(item.get('row_values') or []) for item in failed_rows), default=0))]

    headers = ['Excel工作表', 'Excel行号'] + normalized_headers + ['失败原因']
    widths = [18, 12] + [max(12, min(28, len(header) + 4)) for header in normalized_headers] + [40]

    sheet.append(headers)
    header_fill = PatternFill(fill_type='solid', fgColor='FCE4D6')
    header_font = Font(bold=True)
    for index, _header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = widths[index - 1]

    for item in failed_rows:
        row_values = [_stringify_excel_value(value) for value in (item.get('row_values') or [])]
        padded_values = row_values + [''] * max(0, len(normalized_headers) - len(row_values))
        sheet.append([
            sheet_name,
            item.get('row', ''),
            *padded_values[:len(normalized_headers)],
            item.get('message', ''),
        ])

    sheet.freeze_panes = 'A2'
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _store_import_error_report(sheet_name: str, source_headers, failed_rows):
    if not failed_rows:
        return '', ''

    output = _build_import_error_report(sheet_name, source_headers, failed_rows)
    token = uuid4().hex
    filename = f'合同导入失败明细_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    _IMPORT_ERROR_REPORTS[token] = {
        'filename': filename,
        'content': output.getvalue(),
    }
    return token, filename


def _normalize_excel_option_fields(fields: dict, option_sets: dict) -> dict:
    normalized = dict(fields or {})

    normalized['handling_department'] = _match_option_value(
        normalized.get('handling_department', ''),
        option_sets.get('handling_department', []),
        (normalized.get('handling_department') or '').strip(),
    )
    normalized['project'] = _match_option_value(
        normalized.get('project', ''),
        option_sets.get('project', []),
        (normalized.get('project') or '').strip(),
    )
    normalized['contract_determination_method'] = _match_option_value(
        normalized.get('contract_determination_method', ''),
        option_sets.get('contract_determination_method', []),
        (normalized.get('contract_determination_method') or '').strip(),
    )
    normalized['contract_type'] = _match_option_value(
        normalized.get('contract_type', ''),
        option_sets.get('contract_type', []),
        (normalized.get('contract_type') or '').strip(),
    )
    normalized['purchase_type'] = _match_option_value(
        normalized.get('purchase_type', ''),
        option_sets.get('purchase_type', []),
        (normalized.get('purchase_type') or '').strip(),
    )
    normalized['pricing_method'] = _match_option_value(
        normalized.get('pricing_method', ''),
        option_sets.get('pricing_method', []),
        (normalized.get('pricing_method') or '').strip(),
    )
    normalized['is_archived'] = _match_option_value(
        normalized.get('is_archived', ''),
        option_sets.get('is_archived', []),
        (normalized.get('is_archived') or '').strip(),
    )
    return normalized


def _build_import_payload_from_row(row, field_indexes, header_labels, option_sets):
    payload = {key: '' for key in CONTRACT_FIELD_KEYS}

    for field, index in field_indexes.items():
        raw_value = row[index] if index < len(row) else ''
        if field == 'contract_amount':
            payload[field] = _stringify_excel_value(raw_value)
        elif field == 'handling_date':
            payload[field] = _normalize_excel_date(raw_value)
        else:
            payload[field] = _stringify_excel_value(raw_value)

    payload['contract_name'] = payload['contract_name'].replace('\n', ' ').strip()
    payload['contract_number'] = payload['contract_number'].strip()
    payload['contract_unit'] = payload['contract_unit'].strip()
    payload['copy_count'] = payload.get('copy_count', '').strip()
    payload['save_place'] = payload.get('save_place', '').strip()

    return _normalize_excel_option_fields(payload, option_sets)

def _build_contract_record(body: dict, created_by: str, pending_contract_numbers=None, update_mode=True):
    payload = body or {}
    has_file_path_input = 'file_path' in payload or 'path' in payload
    raw_file_path = payload.get('file_path', payload.get('path'))
    normalized_file_path = _normalize_contract_file_path(raw_file_path) if has_file_path_input else None
    normalized_contract_type = _normalize_contract_type_value(payload.get('contract_type'))
    normalized_stamp_tax_rate = (payload.get('stamp_tax_rate') or '').strip() or STAMP_TAX_RATE_BY_CONTRACT_TYPE.get(normalized_contract_type, '')

    required = ['contract_name', 'handling_department']
    missing = [key for key in required if not str(payload.get(key, '')).strip()]
    if missing:
        return None, f'Missing required fields: {", ".join(missing)}', 400, False

    contract_amount_text = str(payload.get('contract_amount') or '').strip()
    amount = _safe_decimal(contract_amount_text) if contract_amount_text else None
    if contract_amount_text and amount is None:
        return None, 'contract_amount is invalid', 400, False

    copy_count_text = str(payload.get('copy_count') or '').strip()
    if copy_count_text and not re.fullmatch(r'\d+', copy_count_text):
        return None, 'copy_count is invalid', 400, False
    copy_count = int(copy_count_text) if copy_count_text else None

    save_place_text = str(payload.get('save_place') or '').strip()
    if len(save_place_text) > 50:
        return None, 'save_place is too long (max 50)', 400, False
    save_place = save_place_text or None

    contract_number = (payload.get('contract_number') or '').strip()
    if not contract_number:
        return None, '合同编号不能为空', 400, False
    
    pending_contract_numbers = pending_contract_numbers or set()
    is_update = False
    
    if contract_number:
        if contract_number in pending_contract_numbers:
            return None, 'contract_number already exists', 409, False
        
        existing_contract = Contract.query.filter_by(contract_number=contract_number).first()
        if existing_contract:
            # Allow update if contract_number exists but belongs to the same record (e.g. during import with multiple rows sharing the same contract_number)
            if update_mode:
                '''
                if existing_contract.is_archived == '已归档':
                    return None, '已归档的合同只能由管理员进行修改', 403, False
                '''
                existing_contract.contract_name = (payload.get('contract_name') or '').strip()
                existing_contract.contract_unit = (payload.get('contract_unit') or '').strip() or None
                existing_contract.amount = amount
                existing_contract.currency = 'CNY'
                existing_contract.handler = (payload.get('handler') or '').strip() or None
                if has_file_path_input:
                    existing_contract.file_path = normalized_file_path or None
                existing_contract.department = (payload.get('handling_department') or '').strip()
                existing_contract.contract_determination_method = (payload.get('contract_determination_method') or '').strip() or None
                existing_contract.handling_date = _parse_date(payload.get('handling_date'))
                existing_contract.contract_type = normalized_contract_type or None
                existing_contract.purchase_type = (payload.get('purchase_type') or '').strip() or None
                existing_contract.stamp_tax_rate = normalized_stamp_tax_rate or None
                existing_contract.pricing_method = (payload.get('pricing_method') or '').strip() or None
                existing_contract.copy_count = copy_count
                existing_contract.save_place = save_place
                existing_contract.project = (payload.get('project') or '').strip() or None
                existing_contract.fullbody = (payload.get('fullbody') or '').strip() or None
                existing_contract.start_date = _parse_date(payload.get('start_date'))
                existing_contract.end_date = _parse_date(payload.get('end_date'))
                existing_contract.status = (payload.get('status') or 'active').strip() or 'active'
                return existing_contract, '', 0, True
            else:
                return None, 'contract_number already exists', 409, False

    department = (payload.get('handling_department') or '').strip()
    allowed_departments = _get_department_names()
    if department not in allowed_departments:
        return None, 'handling_department is not in configured department settings', 400, False
    _department_dir(department)

    project = (payload.get('project') or '').strip() or None
    if project:
        allowed_projects = _get_project_names()
        if project not in allowed_projects:
            return None, 'project is not in configured project settings', 400, False

    record = Contract(
        contract_number=contract_number or None,
        contract_name=(payload.get('contract_name') or '').strip(),
        contract_unit=(payload.get('contract_unit') or '').strip() or None,
        amount=amount,
        currency='CNY',
        handler=(payload.get('handler') or '').strip() or None,
        department=department,
        contract_determination_method=(payload.get('contract_determination_method') or '').strip() or None,
        handling_date=_parse_date(payload.get('handling_date')),
        contract_type=normalized_contract_type or None,
        purchase_type=(payload.get('purchase_type') or '').strip() or None,
        stamp_tax_rate=normalized_stamp_tax_rate or None,
        pricing_method=(payload.get('pricing_method') or '').strip() or None,
        copy_count=copy_count,
        save_place=save_place,
        is_archived='未归档',
        project=project,
        file_path=normalized_file_path or None,
        fullbody=(payload.get('fullbody') or '').strip() or None,
        start_date=_parse_date(payload.get('start_date')),
        end_date=_parse_date(payload.get('end_date')),
        status=(payload.get('status') or 'active').strip() or 'active',
        created_by=created_by,
    )
    return record, '', 0, False


def _department_dir(department: str) -> str:
    root = current_app.config['CONTRACT_STORAGE_ROOT']
    target = os.path.join(root, department)
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'local':
        os.makedirs(target, exist_ok=True)
    return target


def _build_filestation_path(*parts: str) -> str:
    root = (current_app.config.get('SYNOLOGY_FILESTATION_ROOT') or '/contracts').replace('\\', '/').rstrip('/')
    clean_parts = [str(part).strip('/').replace('\\', '/') for part in parts if str(part).strip('/')]
    if clean_parts:
        return f"{root}/{posixpath.join(*clean_parts)}"
    return root


def _sanitize_upload_filename(file_name: str) -> str:
    base = (file_name or '').strip()
    base = os.path.basename(base).replace('\\', '').replace('/', '')
    if not base:
        return 'upload.bin'
    return base


def _next_available_filename(existing_names, incoming_name: str) -> str:
    existing = set(existing_names)
    if incoming_name not in existing:
        return incoming_name

    stem, ext = os.path.splitext(incoming_name)
    match = re.match(r'^(.*)_(\d+)$', stem)
    if match:
        base_stem = match.group(1)
        number = int(match.group(2)) + 1
    else:
        base_stem = stem
        number = 1

    while True:
        candidate = f"{base_stem}_{number}{ext}"
        if candidate not in existing:
            return candidate
        number += 1


def _synology_upload_login() -> str:
    base_url = current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')
    username = (getattr(g, 'current_user', '') or '').strip()
    if not base_url:
        raise RuntimeError('Missing SYNOLOGY_BASE_URL in .env')
    if not username:
        raise RuntimeError('未找到当前登录用户，请重新登录后重试')

    password = get_cached_user_password(username)
    if not password:
        raise RuntimeError('登录凭据已过期，请重新登录后重试')

    return _synology_user_login(username, password, session_name='FileStation')


def _synology_user_login(account: str, password: str, session_name: str = 'DocsCoolDownload') -> str:
    base_url = current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')
    if not base_url or not account or not password:
        raise RuntimeError('Missing Synology login parameters')

    params = {
        'api': 'SYNO.API.Auth',
        'version': '7',
        'method': 'login',
        'account': account,
        'passwd': password,
        'session': session_name,
        'format': 'sid',
    }

    response = requests.get(
        f"{base_url}/webapi/auth.cgi",
        params=params,
        timeout=EXTERNAL_API_TIMEOUT_SECONDS,
        verify=current_app.config.get('SYNOLOGY_VERIFY_SSL', False),
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get('success'):
        raise RuntimeError(f"Synology 登录失败: {_synology_error_message(payload, 'auth')}")
    return payload.get('data', {}).get('sid', '')


def _synology_api_get(sid: str, params: dict) -> dict:
    base_url = current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')
    endpoint = f"{base_url}/webapi/entry.cgi"
    merged = dict(params)
    merged['_sid'] = sid

    response = requests.get(
        endpoint,
        params=merged,
        timeout=EXTERNAL_API_TIMEOUT_SECONDS,
        verify=current_app.config.get('SYNOLOGY_VERIFY_SSL', False),
    )
    response.raise_for_status()
    return response.json()


def _synology_api_post(sid: str, params: dict, data: dict = None, files: dict = None) -> dict:
    base_url = current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')
    endpoint = f"{base_url}/webapi/entry.cgi"
    merged = dict(params)
    merged['_sid'] = sid

    response = requests.post(
        endpoint,
        params=merged,
        data=data,
        files=files,
        timeout=EXTERNAL_API_TIMEOUT_SECONDS,
        verify=current_app.config.get('SYNOLOGY_VERIFY_SSL', False),
    )
    response.raise_for_status()
    return response.json()


def _synology_json_array(*values: str) -> str:
    return json.dumps([str(value or '') for value in values], ensure_ascii=False)


def _synology_ensure_remote_folder(sid: str, remote_folder: str) -> None:
    parent = posixpath.dirname(remote_folder.rstrip('/')) or '/'
    leaf = posixpath.basename(remote_folder.rstrip('/'))
    if not leaf:
        return

    payload = _synology_api_post(
        sid,
        {
            'api': 'SYNO.FileStation.CreateFolder',
            'version': '2',
            'method': 'create',
        },
        data={
            'folder_path': _synology_json_array(parent),
            'name': _synology_json_array(leaf),
            'force_parent': 'true',
        },
    )

    if payload.get('success'):
        return

    code = _synology_error_code(payload)
    # Existing folder should not block upload.
    if code in {408, 414}:
        return
    raise RuntimeError(f"Synology 创建目录失败: {_synology_error_message(payload, 'filestation')}")


def _synology_list_file_names(sid: str, remote_folder: str):
    payload = _synology_api_get(
        sid,
        {
            'api': 'SYNO.FileStation.List',
            'version': '2',
            'method': 'list',
            'folder_path': remote_folder,
            'additional': '[]',
        },
    )
    if not payload.get('success'):
        raise RuntimeError(f"Synology 目录读取失败: {_synology_error_message(payload, 'filestation')}")

    files = payload.get('data', {}).get('files', [])
    return [item.get('name', '') for item in files if not item.get('isdir')]


def _synology_upload_file(remote_folder: str, file_name: str, uploaded_file) -> None:
    sid = _synology_upload_login()

    _synology_ensure_remote_folder(sid, remote_folder)
    existing_names = _synology_list_file_names(sid, remote_folder)
    resolved_name = _next_available_filename(existing_names, file_name)

    payload = _synology_api_post(
        sid,
        {
            'api': 'SYNO.FileStation.Upload',
            'version': '2',
            'method': 'upload',
        },
        data={
            'path': remote_folder,
            'create_parents': 'true',
            'overwrite': 'false',
        },
        files={
            'file': (resolved_name, uploaded_file.stream, uploaded_file.mimetype or 'application/octet-stream'),
        },
    )
    if not payload.get('success'):
        raise RuntimeError(f"Synology 文件上传失败: {_synology_error_message(payload, 'filestation')}")

    return resolved_name


def _build_synology_file_path(*parts: str) -> str:
    clean_parts = [str(part).strip('/').replace('\\', '/') for part in parts if str(part).strip('/')]
    return posixpath.join(*clean_parts) if clean_parts else ''


def _filename_from_content_disposition(value: str) -> str:
    if not value:
        return ''

    match_utf8 = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.IGNORECASE)
    if match_utf8:
        return unquote(match_utf8.group(1).strip('"'))

    match_plain = re.search(r'filename=([^;]+)', value, flags=re.IGNORECASE)
    if match_plain:
        return match_plain.group(1).strip().strip('"')

    return ''


def _safe_local_file_path(relative_path: str) -> str:
    root = os.path.abspath(current_app.config['CONTRACT_STORAGE_ROOT'])
    normalized = (relative_path or '').replace('\\', '/').lstrip('/')
    resolved = os.path.abspath(os.path.join(root, normalized.replace('/', os.sep)))
    if os.path.commonpath([root, resolved]) != root:
        raise ValueError('invalid file_path')
    return resolved


def _normalize_contract_file_path(raw_path: str) -> str:
    if not raw_path:
        return ''

    value = str(raw_path).strip().replace('\\', '/')
    if not value:
        return ''

    base_url = (current_app.config.get('SYNOLOGY_BASE_URL') or '').strip().rstrip('/')
    storage_root = (current_app.config.get('CONTRACT_STORAGE_ROOT') or '').strip().replace('\\', '/').rstrip('/')
    filestation_root = (current_app.config.get('SYNOLOGY_FILESTATION_ROOT') or '').strip().replace('\\', '/').rstrip('/')

    if base_url and value.startswith(base_url):
        value = value[len(base_url):]
    value = value.lstrip('/')

    if storage_root and value.startswith(storage_root.lstrip('/') + '/'):
        value = value[len(storage_root.lstrip('/')) + 1:]
    elif storage_root and value == storage_root.lstrip('/'):
        value = ''

    if filestation_root and value.startswith(filestation_root.lstrip('/') + '/'):
        value = value[len(filestation_root.lstrip('/')) + 1:]
    elif filestation_root and value == filestation_root.lstrip('/'):
        value = ''

    return value.lstrip('/')


def _load_contract_file_payload(record: Contract):
    normalized_file_path = _normalize_contract_file_path(record.file_path)
    if not normalized_file_path:
        raise FileNotFoundError('该合同未上传文件')

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        password = get_cached_user_password(g.current_user)
        if not password:
            raise PermissionError('登录凭据已过期，请重新登录后下载')

        remote_file_path = _build_filestation_path(normalized_file_path)
        sid = _synology_user_login(g.current_user, password)
        response = None
        for path_value in (f'["{remote_file_path}"]', remote_file_path):
            candidate = requests.get(
                f"{current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')}/webapi/entry.cgi",
                params={
                    'api': 'SYNO.FileStation.Download',
                    'version': '2',
                    'method': 'download',
                    'mode': 'download',
                    'path': path_value,
                    '_sid': sid,
                },
                timeout=EXTERNAL_API_TIMEOUT_SECONDS,
                verify=current_app.config.get('SYNOLOGY_VERIFY_SSL', False),
            )
            if candidate.status_code == 404:
                continue
            candidate.raise_for_status()
            response = candidate
            break

        if response is None:
            raise FileNotFoundError('文件不存在或路径无效')

        content_type = (response.headers.get('Content-Type') or '').lower()
        if 'application/json' in content_type:
            try:
                payload = response.json()
                if not payload.get('success'):
                    raise RuntimeError(_synology_error_message(payload, 'filestation'))
            except ValueError:
                pass

        file_name = _filename_from_content_disposition(response.headers.get('Content-Disposition', ''))
        if not file_name:
            file_name = os.path.basename(normalized_file_path) or f'contract_{record.id}.bin'

        mime = response.headers.get('Content-Type') or 'application/octet-stream'
        return response.content, file_name, mime

    local_file_path = _safe_local_file_path(normalized_file_path)
    if not os.path.isfile(local_file_path):
        raise FileNotFoundError('文件不存在或已被移动')

    file_name = os.path.basename(local_file_path)
    mime = 'application/pdf' if file_name.lower().endswith('.pdf') else 'application/octet-stream'
    with open(local_file_path, 'rb') as f:
        return f.read(), file_name, mime


def _delete_contract_file(record: Contract) -> None:
    normalized_file_path = _normalize_contract_file_path(record.file_path)
    if not normalized_file_path:
        return

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        remote_file_path = _build_filestation_path(normalized_file_path)
        sid = _synology_upload_login()
        payload = _synology_api_post(
            sid,
            {
                'api': 'SYNO.FileStation.Delete',
                'version': '2',
                'method': 'delete',
            },
            data={
                'path': f'["{remote_file_path}"]',
            },
        )
        if payload.get('success'):
            return

        code = _synology_error_code(payload)
        if code in {404, 415}:
            return
        raise RuntimeError(_synology_error_message(payload, 'filestation'))

    local_file_path = _safe_local_file_path(normalized_file_path)
    if os.path.isfile(local_file_path):
        os.remove(local_file_path)


def _normalize_relative_path(raw_path: str) -> str:
    value = (raw_path or '').replace('\\', '/').strip().lstrip('/')
    if not value:
        return ''

    normalized = posixpath.normpath(value)
    if normalized in {'.', ''}:
        return ''
    if normalized == '..' or normalized.startswith('../'):
        raise ValueError('invalid path')
    return normalized.lstrip('/')


def _safe_local_folder_path(relative_path: str) -> str:
    root = os.path.abspath(current_app.config['CONTRACT_STORAGE_ROOT'])
    normalized = _normalize_relative_path(relative_path)
    resolved = os.path.abspath(os.path.join(root, normalized.replace('/', os.sep)))
    if os.path.commonpath([root, resolved]) != root:
        raise ValueError('invalid path')
    return resolved


def _list_local_entries(relative_path: str):
    folder_path = _safe_local_folder_path(relative_path)
    if not os.path.isdir(folder_path):
        raise FileNotFoundError('目录不存在')

    directories = []
    files = []
    with os.scandir(folder_path) as entries:
        for entry in entries:
            entry_rel_path = _build_synology_file_path(relative_path, entry.name)
            if entry.is_dir(follow_symlinks=False):
                directories.append({
                    'name': entry.name,
                    'path': entry_rel_path,
                })
            elif entry.is_file(follow_symlinks=False):
                stat_result = entry.stat(follow_symlinks=False)
                modified_by = getpass.getuser() or '-'
                files.append({
                    'name': entry.name,
                    'path': entry_rel_path,
                    'size': int(stat_result.st_size),
                    'mtime': int(stat_result.st_mtime),
                    'modified_by': modified_by,
                })

    directories.sort(key=lambda item: item['name'].lower())
    files.sort(key=lambda item: item['name'].lower())
    return directories, files


def _remote_folder_path(relative_path: str) -> str:
    normalized = _normalize_relative_path(relative_path)
    return _build_filestation_path(normalized)


def _list_remote_entries(relative_path: str, sid: str = ''):
    resolved_sid = sid or _synology_upload_login()
    folder_path = _remote_folder_path(relative_path)
    payload = _synology_api_get(
        resolved_sid,
        {
            'api': 'SYNO.FileStation.List',
            'version': '2',
            'method': 'list',
            'folder_path': folder_path,
            'additional': '["size","time","owner"]',
        },
    )
    if not payload.get('success'):
        code = _synology_error_code(payload)
        if code in {404, 415}:
            raise FileNotFoundError('目录不存在')
        raise RuntimeError(_synology_error_message(payload, 'filestation'))

    directories = []
    files = []
    for item in payload.get('data', {}).get('files', []):
        name = (item.get('name') or '').strip()
        if not name:
            continue

        entry_rel_path = _build_synology_file_path(relative_path, name)
        if item.get('isdir'):
            directories.append({
                'name': name,
                'path': entry_rel_path,
            })
            continue

        additional = item.get('additional') or {}
        mtime = ((additional.get('time') or {}).get('mtime'))
        size = additional.get('size')
        owner = additional.get('owner') or {}
        modified_by = (
            (owner.get('user') or '').strip()
            or (owner.get('group') or '').strip()
            or (owner.get('uid') or '').strip()
            or '-'
        )
        files.append({
            'name': name,
            'path': entry_rel_path,
            'size': int(size) if isinstance(size, (int, float)) else 0,
            'mtime': int(mtime) if isinstance(mtime, (int, float)) else None,
            'modified_by': modified_by,
        })

    directories.sort(key=lambda entry: entry['name'].lower())
    files.sort(key=lambda entry: entry['name'].lower())
    return directories, files


def _list_storage_entries(relative_path: str):
    normalized = _normalize_relative_path(relative_path)
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        return _list_remote_entries(normalized)
    return _list_local_entries(normalized)


def _list_folder_children_nodes(relative_path: str):
    directories, _files = _list_storage_entries(relative_path)
    return [
        {
            'name': item['name'],
            'path': item['path'],
        }
        for item in directories
    ]


def _count_storage_files_recursive(relative_path: str) -> int:
    normalized = _normalize_relative_path(relative_path)
    queue = [normalized]
    visited = set()
    total = 0

    # Reuse one Synology session for the entire traversal to avoid repeated
    # login/list cycles that may trigger session invalidation (error code 119).
    remote_mode = current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote'
    sid = _synology_upload_login() if remote_mode else ''

    while queue:
        current = queue.pop(0)
        if current in visited:
            continue
        visited.add(current)

        if remote_mode:
            try:
                directories, files = _list_remote_entries(current, sid=sid)
            except RuntimeError as exc:
                # Session may expire during long traversals; re-login once.
                if '错误码: 119' not in str(exc):
                    raise
                sid = _synology_upload_login()
                directories, files = _list_remote_entries(current, sid=sid)
        else:
            directories, files = _list_local_entries(current)

        total += len(files)
        for item in directories:
            child_path = _normalize_relative_path(item.get('path') or '')
            if child_path not in visited:
                queue.append(child_path)

    return total


def _build_folder_tree(relative_path: str):
    directories, _files = _list_storage_entries(relative_path)
    children = [_build_folder_tree(item['path']) for item in directories]
    name = os.path.basename(relative_path.rstrip('/')) if relative_path else ''
    return {
        'name': name or _storage_root_name(),
        'path': relative_path,
        'children': children,
    }


def _storage_root_name() -> str:
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        root = (current_app.config.get('SYNOLOGY_FILESTATION_ROOT') or '').replace('\\', '/').rstrip('/')
    else:
        root = (current_app.config.get('CONTRACT_STORAGE_ROOT') or '').replace('\\', '/').rstrip('/')
    return os.path.basename(root) or root or '/'


def _build_contract_file_index() -> dict:
    index = {}
    rows = Contract.query.filter(Contract.file_path.isnot(None)).all()
    for row in rows:
        normalized = _normalize_contract_file_path(row.file_path)
        if not normalized:
            continue
        if normalized not in index:
            index[normalized] = row
    return index


def _build_folder_file_items(relative_folder_path: str):
    _directories, files = _list_storage_entries(relative_folder_path)
    contract_index = _build_contract_file_index()

    payload = []
    for item in files:
        relative_file_path = item['path']
        matched = contract_index.get(relative_file_path)
        contract_payload = matched.to_dict() if matched else None

        row = {
            'name': item['name'],
            'file_path': relative_file_path,
            'size': item.get('size') or 0,
            'mtime': item.get('mtime'),
            'matched_contract_id': matched.id if matched else None,
            'contract_name': contract_payload.get('contract_name') if contract_payload else '<无匹配>',
            'contract_number': contract_payload.get('contract_number') if contract_payload else '',
            'contract_unit': contract_payload.get('contract_unit') if contract_payload else '',
            'contract_amount': contract_payload.get('contract_amount') if contract_payload else '',
            'handler': contract_payload.get('handler') if contract_payload else '',
            'handling_department': contract_payload.get('handling_department') if contract_payload else '',
            'handling_date': contract_payload.get('handling_date') if contract_payload else '',
            'contract_type': contract_payload.get('contract_type') if contract_payload else '',
            'purchase_type': contract_payload.get('purchase_type') if contract_payload else '',
            'stamp_tax_rate': contract_payload.get('stamp_tax_rate') if contract_payload else '',
            'copy_count': contract_payload.get('copy_count') if contract_payload else None,
            'save_place': contract_payload.get('save_place') if contract_payload else '',
            'is_archived': contract_payload.get('is_archived') if contract_payload else '',
            'project': contract_payload.get('project') if contract_payload else '',
            'contract': contract_payload,
        }
        payload.append(row)

    return payload


def _normalize_match_text(value: str) -> str:
    text = str(value or '').strip().lower()
    if not text:
        return ''
    text = re.sub(r'\.(pdf|PDF)$', '', text, flags=re.IGNORECASE)
    text = re.sub(r'[^\w\u4e00-\u9fff]+', '', text)
    return text


def _extract_match_key_from_filename(file_name: str) -> str:
    base_name = os.path.splitext(os.path.basename(str(file_name or '')))[0]
    if not base_name:
        return ''

    chinese_indexes = [idx for idx, ch in enumerate(base_name) if re.match(r'[\u4e00-\u9fff]', ch)]
    if not chinese_indexes:
        return ''

    start_idx = chinese_indexes[0]
    end_idx = chinese_indexes[-1]
    return base_name[start_idx:end_idx + 1].strip()


def _select_best_contract_by_key(match_key: str, candidates: list):
    normalized_key = _normalize_match_text(match_key)
    if not normalized_key:
        return None, '', []

    scored = []
    for row in candidates:
        normalized_contract_name = _normalize_match_text(row.contract_name)
        if not normalized_contract_name:
            continue

        exact = normalized_contract_name == normalized_key
        contains = normalized_key in normalized_contract_name
        if not exact and not contains:
            continue

        similarity = SequenceMatcher(None, normalized_key, normalized_contract_name).ratio()
        scored.append({
            'row': row,
            'normalized_name': normalized_contract_name,
            'exact': exact,
            'contains': contains,
            'similarity': similarity,
        })

    if not scored:
        return None, '', []

    exact_rows = [item for item in scored if item['exact']]
    if exact_rows:
        exact_rows.sort(key=lambda item: (-item['similarity'], item['row'].id))
        return exact_rows[0]['row'], 'exact', exact_rows

    contains_rows = [item for item in scored if item['contains']]
    if len(contains_rows) == 1:
        return contains_rows[0]['row'], 'contains-single', contains_rows

    contains_rows.sort(key=lambda item: (-item['similarity'], item['row'].id))
    return contains_rows[0]['row'], 'contains-best', contains_rows


def _collect_storage_pdf_files() -> list:
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        stack = ['']
        result = []
        while stack:
            current_path = stack.pop()
            directories, files = _list_remote_entries(current_path, sid=sid)
            for item in files:
                name = item.get('name') or ''
                if name.lower().endswith('.pdf'):
                    result.append({
                        'name': name,
                        'path': item.get('path') or '',
                        'mtime': item.get('mtime') or 0,
                        'modified_by': (item.get('modified_by') or '').strip() or '-',
                    })
            for directory in directories:
                child_path = directory.get('path') or ''
                if child_path:
                    stack.append(child_path)
        return result

    root = os.path.abspath(current_app.config['CONTRACT_STORAGE_ROOT'])
    result = []
    for dirpath, _dirnames, filenames in os.walk(root):
        for filename in filenames:
            if not filename.lower().endswith('.pdf'):
                continue
            full_path = os.path.join(dirpath, filename)
            relative_path = os.path.relpath(full_path, root).replace('\\', '/')
            try:
                mtime = int(os.path.getmtime(full_path))
            except OSError:
                mtime = 0
            result.append({
                'name': filename,
                'path': relative_path,
                'mtime': mtime,
                'modified_by': getpass.getuser() or '-',
            })
    return result


def _collect_storage_files() -> list:
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        stack = ['']
        result = []
        while stack:
            current_path = stack.pop()
            directories, files = _list_remote_entries(current_path, sid=sid)
            for item in files:
                result.append({
                    'name': item.get('name') or '',
                    'path': item.get('path') or '',
                    'mtime': item.get('mtime') or 0,
                })
            for directory in directories:
                child_path = directory.get('path') or ''
                if child_path:
                    stack.append(child_path)
        return result

    root = os.path.abspath(current_app.config['CONTRACT_STORAGE_ROOT'])
    result = []
    for dirpath, _dirnames, filenames in os.walk(root):
        for filename in filenames:
            full_path = os.path.join(dirpath, filename)
            relative_path = os.path.relpath(full_path, root).replace('\\', '/')
            try:
                mtime = int(os.path.getmtime(full_path))
            except OSError:
                mtime = 0
            result.append({
                'name': filename,
                'path': relative_path,
                'mtime': mtime,
            })
    return result


def _select_best_pdf_match(contract_name: str, pdf_files: list):
    normalized_contract = _normalize_match_text(contract_name)
    if not normalized_contract:
        return None, []

    matched = []
    for item in pdf_files:
        file_name = item.get('name') or ''
        normalized_file = _normalize_match_text(file_name)
        if not normalized_file:
            continue
        if normalized_contract in normalized_file:
            similarity = SequenceMatcher(None, normalized_contract, normalized_file).ratio()
            matched.append({
                'name': file_name,
                'path': item.get('path') or '',
                'similarity': similarity,
                'mtime': item.get('mtime') or 0,
            })

    if not matched:
        return None, []

    matched.sort(key=lambda row: (-row['similarity'], -row['mtime'], row['name']))
    return matched[0], matched


def _create_storage_folder(parent_path: str, folder_name: str) -> str:
    normalized_parent = _normalize_relative_path(parent_path)
    name = (folder_name or '').strip()
    if not name:
        raise ValueError('文件夹名称不能为空')
    if '/' in name or '\\' in name:
        raise ValueError('文件夹名称不能包含斜杠')

    target_relative_path = _build_synology_file_path(normalized_parent, name)
    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        payload = _synology_api_post(
            sid,
            {
                'api': 'SYNO.FileStation.CreateFolder',
                'version': '2',
                'method': 'create',
            },
            data={
                'folder_path': _synology_json_array(_remote_folder_path(normalized_parent)),
                'name': _synology_json_array(name),
                'force_parent': 'false',
            },
        )
        if not payload.get('success'):
            code = _synology_error_code(payload)
            if code == 414:
                raise FileExistsError('文件夹已存在')
            raise RuntimeError(_synology_error_message(payload, 'filestation'))
        return target_relative_path

    target_path = _safe_local_folder_path(target_relative_path)
    os.makedirs(target_path, exist_ok=False)
    return target_relative_path


def _delete_storage_folder(relative_path: str) -> None:
    normalized = _normalize_relative_path(relative_path)
    if not normalized:
        raise ValueError('不允许删除根目录')

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        directories, files = _list_remote_entries(normalized, sid=sid)
        if files:
            raise RuntimeError('该文件夹下存在文件，不能删除')
        if directories:
            raise RuntimeError('该文件夹下存在子文件夹，不能删除')

        payload = _synology_api_post(
            sid,
            {
                'api': 'SYNO.FileStation.Delete',
                'version': '2',
                'method': 'delete',
            },
            data={
                'path': f'["{_remote_folder_path(normalized)}"]',
                'recursive': 'false',
            },
        )
        if not payload.get('success'):
            code = _synology_error_code(payload)
            if code in {404, 415}:
                raise FileNotFoundError('目录不存在')
            raise RuntimeError(_synology_error_message(payload, 'filestation'))
        return

    folder_path = _safe_local_folder_path(normalized)
    if not os.path.isdir(folder_path):
        raise FileNotFoundError('目录不存在')

    child_dirs = []
    child_files = []
    with os.scandir(folder_path) as entries:
        for entry in entries:
            if entry.is_file(follow_symlinks=False):
                child_files.append(entry.name)
            elif entry.is_dir(follow_symlinks=False):
                child_dirs.append(entry.name)

    if child_files:
        raise RuntimeError('该文件夹下存在文件，不能删除')
    if child_dirs:
        raise RuntimeError('该文件夹下存在子文件夹，不能删除')

    os.rmdir(folder_path)


def _load_storage_file_payload(relative_file_path: str):
    normalized = _normalize_relative_path(relative_file_path)
    if not normalized:
        raise ValueError('file_path is required')

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        remote_file_path = _remote_folder_path(normalized)
        response = None
        for path_value in (f'["{remote_file_path}"]', remote_file_path):
            candidate = requests.get(
                f"{current_app.config.get('SYNOLOGY_BASE_URL', '').rstrip('/')}/webapi/entry.cgi",
                params={
                    'api': 'SYNO.FileStation.Download',
                    'version': '2',
                    'method': 'download',
                    'mode': 'download',
                    'path': path_value,
                    '_sid': sid,
                },
                timeout=EXTERNAL_API_TIMEOUT_SECONDS,
                verify=current_app.config.get('SYNOLOGY_VERIFY_SSL', False),
            )
            if candidate.status_code == 404:
                continue
            candidate.raise_for_status()
            response = candidate
            break

        if response is None:
            raise FileNotFoundError('文件不存在或路径无效')

        content_type = (response.headers.get('Content-Type') or '').lower()
        if 'application/json' in content_type:
            try:
                payload = response.json()
                if not payload.get('success'):
                    raise RuntimeError(_synology_error_message(payload, 'filestation'))
            except ValueError:
                pass

        file_name = _filename_from_content_disposition(response.headers.get('Content-Disposition', ''))
        if not file_name:
            file_name = os.path.basename(normalized) or 'download.bin'

        mime = response.headers.get('Content-Type') or mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
        return response.content, file_name, mime

    local_file_path = _safe_local_file_path(normalized)
    if not os.path.isfile(local_file_path):
        raise FileNotFoundError('文件不存在或已被移动')

    file_name = os.path.basename(local_file_path)
    mime = mimetypes.guess_type(file_name)[0] or 'application/octet-stream'
    with open(local_file_path, 'rb') as f:
        return f.read(), file_name, mime


def _delete_storage_file(relative_file_path: str) -> str:
    normalized = _normalize_relative_path(relative_file_path)
    if not normalized:
        raise ValueError('file_path is required')

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        remote_file_path = _remote_folder_path(normalized)
        sid = _synology_upload_login()
        payload = _synology_api_post(
            sid,
            {
                'api': 'SYNO.FileStation.Delete',
                'version': '2',
                'method': 'delete',
            },
            data={
                'path': f'["{remote_file_path}"]',
            },
        )
        if not payload.get('success'):
            code = _synology_error_code(payload)
            if code in {404, 415}:
                raise FileNotFoundError('文件不存在')
            raise RuntimeError(_synology_error_message(payload, 'filestation'))
        return normalized

    local_file_path = _safe_local_file_path(normalized)
    if not os.path.isfile(local_file_path):
        raise FileNotFoundError('文件不存在')

    os.remove(local_file_path)
    return normalized


def _rename_storage_file(relative_file_path: str, new_name: str) -> tuple[str, str]:
    normalized = _normalize_relative_path(relative_file_path)
    if not normalized:
        raise ValueError('file_path is required')

    target_name = (new_name or '').strip()
    if not target_name:
        raise ValueError('文件名不能为空')
    if '/' in target_name or '\\' in target_name:
        raise ValueError('文件名不能包含斜杠')

    parent_path = posixpath.dirname(normalized)
    if parent_path == '.':
        parent_path = ''
    new_relative_path = _build_synology_file_path(parent_path, target_name)

    if new_relative_path == normalized:
        return normalized, normalized

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        sid = _synology_upload_login()
        old_remote_path = _remote_folder_path(normalized)
        payload = _synology_api_post(
            sid,
            {
                'api': 'SYNO.FileStation.Rename',
                'version': '2',
                'method': 'rename',
            },
            data={
                'path': f'["{old_remote_path}"]',
                'name': _synology_json_array(target_name),
            },
        )
        if not payload.get('success'):
            code = _synology_error_code(payload)
            if code in {404, 415}:
                raise FileNotFoundError('文件不存在')
            if code in {405, 408}:
                raise FileExistsError('同名文件已存在')
            raise RuntimeError(_synology_error_message(payload, 'filestation'))
        return normalized, new_relative_path

    old_local_path = _safe_local_file_path(normalized)
    if not os.path.isfile(old_local_path):
        raise FileNotFoundError('文件不存在')

    new_local_path = _safe_local_file_path(new_relative_path)
    if os.path.exists(new_local_path):
        raise FileExistsError('同名文件已存在')

    os.rename(old_local_path, new_local_path)
    return normalized, new_relative_path


def _clear_contract_file_path_by_relative_path(relative_file_path: str) -> list[int]:
    normalized_target = _normalize_relative_path(relative_file_path)
    if not normalized_target:
        return []

    affected_ids = []
    rows = Contract.query.filter(Contract.file_path.isnot(None)).all()
    for row in rows:
        if _normalize_contract_file_path(row.file_path) != normalized_target:
            continue
        row.file_path = None
        affected_ids.append(row.id)

    return affected_ids


def _replace_contract_file_path_by_relative_path(old_relative_path: str, new_relative_path: str) -> list[int]:
    normalized_old = _normalize_relative_path(old_relative_path)
    normalized_new = _normalize_relative_path(new_relative_path)
    if not normalized_old or not normalized_new:
        return []

    affected_ids = []
    rows = Contract.query.filter(Contract.file_path.isnot(None)).all()
    for row in rows:
        if _normalize_contract_file_path(row.file_path) != normalized_old:
            continue
        row.file_path = normalized_new
        affected_ids.append(row.id)

    return affected_ids


def _get_department_names():
    rows = Department.query.order_by(Department.name.asc()).all()
    return [row.name for row in rows]


def _get_project_names():
    rows = ProjectOption.query.order_by(ProjectOption.name.asc()).all()
    return [row.name for row in rows]


def _merge_options(default_values, db_values):
    merged = []
    seen = set()
    for item in default_values + db_values:
        value = (item or '').strip()
        if not value or value in seen:
            continue
        seen.add(value)
        merged.append(value)
    return merged


def _synology_error_code(payload: dict):
    error = payload.get('error') or {}
    code = error.get('code')
    if code is None:
        return None
    try:
        return int(code)
    except (TypeError, ValueError):
        return None


def _synology_error_message(payload: dict, scope: str) -> str:
    code = _synology_error_code(payload)
    if code is None:
        return 'Synology 未返回具体错误码'

    if scope == 'auth':
        text = SYNOLOGY_AUTH_ERROR_MESSAGES.get(code)
    elif scope == 'filestation':
        text = SYNOLOGY_FILESTATION_ERROR_MESSAGES.get(code)
    else:
        text = None

    if text is None:
        text = SYNOLOGY_COMMON_ERROR_MESSAGES.get(code, '未知错误')
    return f'{text}(错误码: {code})'


def _preview_lines(text: str, limit: int = 6):
    lines = [line.strip() for line in (text or '').splitlines() if line.strip()]
    return lines[:limit]


## OCR相关函数已迁移到ocr_utils.py



## OCR相关函数已迁移到ocr_utils.py



## OCR相关函数已迁移到ocr_utils.py



## OCR相关函数已迁移到ocr_utils.py

def _extract_ai_content(payload: dict) -> str:
    if not isinstance(payload, dict):
        return ''

    output_text = payload.get('output_text')
    if isinstance(output_text, str) and output_text.strip():
        return output_text

    reply = payload.get('reply')
    if isinstance(reply, str):
        return reply

    if isinstance(payload.get('message'), str):
        return payload.get('message')

    choices = payload.get('choices')
    if isinstance(choices, list) and choices:
        message = choices[0].get('message') if isinstance(choices[0], dict) else None
        if isinstance(message, dict):
            content = message.get('content')
            if isinstance(content, str):
                return content
            if isinstance(content, list):
                text_parts = []
                for item in content:
                    if isinstance(item, dict) and isinstance(item.get('text'), str):
                        text_parts.append(item.get('text'))
                if text_parts:
                    return '\n'.join(text_parts)
        text = choices[0].get('text') if isinstance(choices[0], dict) else None
        if isinstance(text, str) and text.strip():
            return text

    data = payload.get('data')
    if isinstance(data, dict) and isinstance(data.get('reply'), str):
        return data.get('reply')

    if isinstance(data, dict):
        nested_choices = data.get('choices')
        if isinstance(nested_choices, list) and nested_choices:
            nested_msg = nested_choices[0].get('message') if isinstance(nested_choices[0], dict) else None
            if isinstance(nested_msg, dict):
                nested_content = nested_msg.get('content')
                if isinstance(nested_content, str) and nested_content.strip():
                    return nested_content

    result = payload.get('result')
    if isinstance(result, str) and result.strip():
        return result

    return ''


def _extract_json_object(text: str) -> dict:
    if not text:
        return {}

    stripped = text.strip()
    try:
        parsed = json.loads(stripped)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        pass

    start = stripped.find('{')
    end = stripped.rfind('}')
    if start >= 0 and end > start:
        candidate = stripped[start:end + 1]
        try:
            parsed = json.loads(candidate)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}

    return {}


def _normalize_date_value(value: str) -> str:
    text = (value or '').strip()
    if not text:
        return ''

    match = re.search(r'(20\d{2})[-/.年\s]+(\d{1,2})[-/.月\s]+(\d{1,2})', text)
    if not match:
        return text

    year, month, day = match.groups()
    return f'{year}-{int(month):02d}-{int(day):02d}'


def _find_contract_number(pdf_text: str, fallback: str) -> str:
    text = (pdf_text or '').replace(' ', '')
    patterns = [
        r'合同编号[:：]?([A-Za-z]{1,4}\d{2,8}-\d{1,4})',
        r'\b([A-Za-z]{1,4}\d{2,8}-\d{1,4})\b',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1).upper()

    cleaned = re.sub(r'[^A-Za-z0-9-]', '', (fallback or '').upper())
    return cleaned


def _find_amount(fallback: str) -> str:
    # 优先解析“数字+单位”（亿/万/千/元），并统一换算成元，最多保留到分。
    text = str(fallback or '').replace(',', '').replace('，', '').strip()
    if not text:
        return ''

    unit_match = re.search(r'([+-]?[0-9]+(?:\.[0-9]+)?)(亿元|万元|千元|万|亿|元)', text)
    if unit_match:
        number_text, unit_text = unit_match.group(1), unit_match.group(2)
        try:
            amount_yuan = Decimal(number_text) * AI_AMOUNT_UNIT_TO_YUAN[unit_text]
            amount_yuan = amount_yuan.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return _format_decimal_plain(amount_yuan)
        except (InvalidOperation, ValueError, KeyError):
            return ''

    # 无单位时按“元”处理，最多保留到分。
    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', text):
        try:
            amount_yuan = Decimal(text).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return _format_decimal_plain(amount_yuan)
        except (InvalidOperation, ValueError):
            return ''

    # 兜底清洗，处理如“¥1,234.5元”这类混合字符串。
    cleaned = re.sub(r'[^0-9.+-]', '', text)
    if cleaned.count('.') > 1:
        first_dot = cleaned.find('.')
        cleaned = cleaned[:first_dot + 1] + cleaned[first_dot + 1:].replace('.', '')

    if re.fullmatch(r'[+-]?[0-9]+(?:\.[0-9]+)?', cleaned):
        try:
            amount_yuan = Decimal(cleaned).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return _format_decimal_plain(amount_yuan)
        except (InvalidOperation, ValueError):
            return ''

    return ''


def _normalize_company_name(value: str) -> str:
    text = (value or '').strip()
    if not text:
        return ''
    return re.sub(r'[^\w\u4e00-\u9fff]+', '', text).lower()


def _exclude_my_company(contract_unit: str) -> str:
    my_comp = current_app.config.get('MY_COMP', '')
    if not my_comp:
        return (contract_unit or '').strip()

    parts = [
        part.strip()
        for part in re.split(r'[、,，;；/\\\n]+', contract_unit or '')
        if part and part.strip()
    ]
    if not parts:
        return ''

    my_comp_norm = _normalize_company_name(my_comp)
    kept = [part for part in parts if _normalize_company_name(part) != my_comp_norm]

    if not kept:
        return ''
    return '、'.join(kept)


def _normalize_ai_fields(raw: dict, pdf_text: str = '') -> dict:
    normalized = {}
    for key in CONTRACT_FIELD_KEYS:
        value = raw.get(key) if isinstance(raw, dict) else None
        if value is None:
            normalized[key] = ''
        elif isinstance(value, (int, float)):
            normalized[key] = str(value)
        else:
            normalized[key] = str(value).strip()

    normalized['contract_number'] = _find_contract_number(pdf_text, normalized.get('contract_number', ''))
    normalized['contract_unit'] = _exclude_my_company(normalized.get('contract_unit', ''))
    
    normalized['contract_amount'] = _find_amount(normalized.get('contract_amount', ''))

    normalized['handling_date'] = _normalize_date_value(normalized.get('handling_date', ''))
    
    return normalized


def _normalize_option_text(value: str) -> str:
    text = (value or '').strip().lower()
    return re.sub(r'[^\w\u4e00-\u9fff]+', '', text)


def _find_ai_match_candidates(fields: dict, limit: int = AI_MATCH_CANDIDATE_LIMIT):
    normalized_name = _normalize_option_text((fields or {}).get('contract_name', ''))
    amount = _safe_decimal((fields or {}).get('contract_amount'))
    if not normalized_name and amount is None:
        return []

    same_amount_candidates = []
    name_similarity_candidates = []
    rows = Contract.query.order_by(Contract.updated_at.desc()).all()
    for row in rows:
        reasons = []
        similarity = 0.0
        contains_match = False

        existing_name = _normalize_option_text(row.contract_name or '')
        if normalized_name and existing_name:
            similarity = SequenceMatcher(None, normalized_name, existing_name).ratio()
            contains_match = normalized_name in existing_name or existing_name in normalized_name
            if contains_match or similarity >= 0.55:
                reasons.append('标题相似')

        same_amount = False
        if amount is not None and row.amount is not None:
            try:
                same_amount = Decimal(str(row.amount)) == amount
            except (InvalidOperation, ValueError):
                same_amount = False
            if same_amount:
                reasons.append('金额相同')

        if not reasons:
            continue

        item = row.to_dict()
        item['match_reasons'] = reasons
        item['name_similarity'] = round(similarity, 4)
        item['same_amount'] = same_amount

        if same_amount:
            same_amount_candidates.append(item)
        elif contains_match or similarity >= 0.55:
            name_similarity_candidates.append(item)

    same_amount_candidates.sort(key=lambda item: (-item['name_similarity'], -item['id']))
    name_similarity_candidates.sort(key=lambda item: (-item['name_similarity'], -item['id']))

    merged = []
    seen_ids = set()
    for item in same_amount_candidates:
        if item['id'] in seen_ids:
            continue
        merged.append(item)
        seen_ids.add(item['id'])

    for item in name_similarity_candidates[:5]:
        if item['id'] in seen_ids:
            continue
        merged.append(item)
        seen_ids.add(item['id'])

    return merged[:limit]


def _match_option_value(value: str, options, default: str = '') -> str:
    candidates = [item for item in (options or []) if str(item).strip()]
    if not candidates:
        return default

    raw = (value or '').strip()
    if not raw:
        return default

    if raw in candidates:
        return raw

    norm_raw = _normalize_option_text(raw)
    if not norm_raw:
        return default

    for item in candidates:
        if _normalize_option_text(item) == norm_raw:
            return item

    best = ''
    best_score = 0
    for item in candidates:
        norm_item = _normalize_option_text(item)
        if not norm_item:
            continue
        if norm_item in norm_raw or norm_raw in norm_item:
            score = min(len(norm_item), len(norm_raw))
            if score > best_score:
                best = item
                best_score = score

    if best:
        return best
    return default


def _get_contract_option_sets() -> dict:
    option_sets = {
        key: list(CSV_OPTION_DEFAULTS.get(key, []))
        for key in CSV_OPTION_DEFAULTS.keys()
    }
    option_sets['handling_department'] = _get_department_names()
    option_sets['project'] = _merge_options(_get_project_names(), [OPTION_FIELD_DEFAULTS['project']])
    return option_sets


def _normalize_option_fields(fields: dict) -> dict:
    option_sets = _get_contract_option_sets()
    normalized = dict(fields or {})

    normalized['handling_department'] = _match_option_value(
        normalized.get('handling_department', ''),
        option_sets.get('handling_department', []),
        '',
    )
    normalized['project'] = _match_option_value(
        normalized.get('project', ''),
        option_sets.get('project', []),
        OPTION_FIELD_DEFAULTS['project'],
    )


    normalized['contract_determination_method'] = _match_option_value(
        normalized.get('contract_determination_method', ''),
        option_sets.get('contract_determination_method', []),
        OPTION_FIELD_DEFAULTS['contract_determination_method'],
    )


    normalized['contract_type'] = _match_option_value(
        normalized.get('contract_type', ''),
        option_sets.get('contract_type', []),
        '',
    )


    normalized['purchase_type'] = _match_option_value(
        normalized.get('purchase_type', ''),
        option_sets.get('purchase_type', []),
        '',
    )
    normalized['pricing_method'] = _match_option_value(
        normalized.get('pricing_method', ''),
        option_sets.get('pricing_method', []),
        '',
    )
    normalized['is_archived'] = _match_option_value(
        normalized.get('is_archived', ''),
        option_sets.get('is_archived', []),
        OPTION_FIELD_DEFAULTS['is_archived'],
    )
    return normalized


def _has_any_field_value(fields: dict) -> bool:
    if not isinstance(fields, dict):
        return False
    return any(str(fields.get(key, '')).strip() for key in CONTRACT_FIELD_KEYS)


def _minimax_extract_fields(pdf_text: str) -> dict:

    #print('pdf_text:', pdf_text[:2000])

    api_key = (current_app.config.get('MINIMAX_API_KEY') or '').strip()
    if not api_key:
        raise RuntimeError('MINIMAX_API_KEY 未配置')

    api_url = (current_app.config.get('MINIMAX_API_URL') or '').strip()
    model = (current_app.config.get('MINIMAX_MODEL') or '').strip()

    prompt = (
        'contract_type指的是合同类型，值在如下选择：' + ','.join(CSV_OPTION_DEFAULTS['contract_type']) + '。必须要返回内容，如果文本没有明确写出类型，请按印花税合同分类选择最贴近的一项：设备材料采购/销售归为买卖合同，贷款融资归为借款合同，房屋设备租用归为租赁合同，委托加工制作归为承揽合同，施工建设归为建设工程合同，货物物流承运归为运输合同，技术开发转让咨询服务归为技术合同，代保管归为保管合同，仓储服务归为仓储合同，保险保单归为财产保险合同。\n'
        'purchase_type指的是采购类型，值在如下选择：' + ','.join(CSV_OPTION_DEFAULTS['purchase_type']) + '。必须要返回内容，按合同业务性质归类：工程施工建设归工程类，咨询运维检测培训等归服务类，设备材料货物采购归采购类，不属于采购项目或与采购无关归非采购类。\n'
        'stamp_tax_rate指的是印花税率，请根据合同类型返回税法规定税率：买卖合同0.03%，借款合同0.005%，租赁合同0.1%，承揽合同0.03%，建设工程合同0.03%，运输合同0.03%，技术合同0.03%，保管合同0.1%，仓储合同0.1%，财产保险合同0.1%，其他类型可返回空字符串。\n'
        'pricing_method指的是合同的计价方式，值在如下选择：' + ','.join(CSV_OPTION_DEFAULTS['pricing_method']) + '。必须要返回内容，如果找不到就请总结提炼这个合同可能是通过什么方式计价的，如果讲到了综合单价暂定工程量就是单价合同，其它默认都返回总价合同。\n'
        'contract_determination_method指的是合同是如何确定的，值在如下选择：' + ','.join(CSV_OPTION_DEFAULTS['contract_determination_method']) + '，必须要返回内容，如果找不到就请总结提炼这个合同可能是通过什么方式确定的。\n'
        'contract_name指的是合同名称或标题，必须要返回内容，一般会出现在文本的前几行，如果找不到就请总结合同标题， 如果某字段找不到准确的文本，请尽量根据上下文来总结提炼。\n'
        'handling_date格式为 YYYY-MM-DD。\n'
        'contract_unit指的是对方的公司，因此不能返回我方公司名称“' + (current_app.config.get('MY_COMP') or '') + '”及其常见变体，如果不好定位就返回文本里出现的非我公司的单位名称。\n'
        'contract_amount 指的是合同金额（人民币元）返回以为元单位的纯数字字符串，如果合同文本中带单位，如果原文是万元、亿元等其它单位返回时带上万、亿\n'
        'project是合同属于什么工程或项目，请尽量从标题或是其它文本中识别出项目相关信息， 从全文解理本合同是不是属于如下项目列表，不需要要强匹配找意思相似的标题或文本, 必须值返回如下的项目名称文本，如果合同真的不属于项目或是工程请返回空""：'+ ','.join(_get_project_names()) + '。\n' 
        'handling_department必须返回如下的部门列表其中之一文本（如果能从标题或是其它文本中识别出部门相关信息的话最好，不能的话先判断这个合同一般是列表中的哪个部门职责，通过判断来返回），实在靠不上部门请返回空""：' + ','.join(_get_department_names()) + '。\n'
        '以下是PDF文本：\n'
        + pdf_text[:20000]
    )

    payload = {
        'model': model,
        'messages': [
            {
                'role': 'system',
                'content': '你是PDF扫描合同得出文本的结构化抽取专家，请从给定PDF扫描出来的合同文本，该合同文本是从图像识别出来没有经过加工肯定包含意外的回车、噪声、格式和语序问题。只抽取合同字段，并只返回JSON对象，不要输出任何解释。你是合同结构化抽取助手，只返回合法JSON对象，不输出解释。'
                '返回的JSON键必须严格使用以下字段：'
                +','.join(CONTRACT_FIELD_KEYS)
                + '\n'
                   
            },
            {
                'role': 'user',
                'content': prompt,
            },
        ],
        'temperature': 0.2,
        'max_tokens': 1024,
        'stream': False,
    }


    #print('payload:',payload)

    current_app.logger.info(
        'AI parse: Minimax request model=%s text_chars=%s prompt_chars=%s',
        model,
        len(pdf_text),
        len(prompt),
    )

    response = requests.post(
        api_url,
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        },
        json=payload,
        timeout=EXTERNAL_API_TIMEOUT_SECONDS,
    )
    current_app.logger.info('AI parse: Minimax response status=%s', response.status_code)
    response.raise_for_status()

    response_payload = response.json()
    current_app.logger.info('AI parse: Minimax payload keys=%s', sorted(list(response_payload.keys())))
    base_resp = response_payload.get('base_resp')
    
    if isinstance(base_resp, dict):
        current_app.logger.info('AI parse: Minimax base_resp=%s', base_resp)
        status_code = base_resp.get('status_code')
        if status_code not in (0, None):
            raise RuntimeError(f"Minimax接口错误: {base_resp.get('status_msg') or status_code}")

    content = _extract_ai_content(response_payload)


    current_app.logger.info('AI parse: Minimax content chars=%s snippet=%s', len(content), (content or '')[:200])
    if not content.strip():
        raise RuntimeError('Minimax返回成功但无可用文本内容')

    parsed = _extract_json_object(content)
    current_app.logger.info('AI parse: Minimax parsed keys=%s', sorted(list(parsed.keys())) if isinstance(parsed, dict) else [])
    return _normalize_ai_fields(parsed, pdf_text)


## OCR相关函数已迁移到ocr_utils.py



## OCR相关函数已迁移到ocr_utils.py



@contracts_bp.get('/departments')
@require_auth
def list_departments():
    return jsonify(_get_department_names())


@contracts_bp.get('/settings/departments')
@require_auth
def list_department_settings():
    rows = Department.query.order_by(Department.name.asc()).all()
    return jsonify([row.to_dict() for row in rows])


@contracts_bp.get('/settings/projects')
@require_auth
def list_project_settings():
    rows = ProjectOption.query.order_by(ProjectOption.name.asc()).all()
    return jsonify([row.to_dict() for row in rows])


@contracts_bp.post('/settings/projects')
@require_auth
def create_project_setting():
    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()

    if not name:
        return jsonify({'message': 'name is required'}), 400
    if len(name) > 255:
        return jsonify({'message': '项目名称最多255个字符'}), 400
    if ProjectOption.query.filter_by(name=name).first():
        return jsonify({'message': '项目已存在'}), 409

    row = ProjectOption(name=name)
    db.session.add(row)
    db.session.commit()
    return jsonify(row.to_dict()), 201


@contracts_bp.delete('/settings/projects/<int:project_id>')
@require_auth
def delete_project_setting(project_id):
    row = ProjectOption.query.get_or_404(project_id)

    in_use = Contract.query.filter(Contract.project == row.name).first()
    if in_use:
        return jsonify({'message': '该项目下已有合同，无法删除'}), 409

    db.session.delete(row)
    db.session.commit()
    return jsonify({'success': True})


@contracts_bp.get('/options/contract-fields')
@require_auth
def contract_field_options():
    payload = {
        key: list(CSV_OPTION_DEFAULTS.get(key, []))
        for key in CSV_OPTION_DEFAULTS.keys()
    }
    payload['project'] = _get_project_names()
    payload['stamp_tax_rate_by_contract_type'] = STAMP_TAX_RATE_BY_CONTRACT_TYPE
    return jsonify(payload)


@contracts_bp.post('/settings/departments')
@require_auth
def create_department_setting():
    body = request.get_json(silent=True) or {}
    name = (body.get('name') or '').strip()

    if not name:
        return jsonify({'message': 'name is required'}), 400
    if len(name) > 50:
        return jsonify({'message': '部门名称最多50个字符'}), 400
    if Department.query.filter_by(name=name).first():
        return jsonify({'message': '部门已存在'}), 409

    row = Department(name=name)
    db.session.add(row)
    db.session.commit()

    _department_dir(name)
    return jsonify(row.to_dict()), 201


@contracts_bp.delete('/settings/departments/<int:department_id>')
@require_auth
def delete_department_setting(department_id):
    row = Department.query.get_or_404(department_id)

    if row.name == DEFAULT_DEPARTMENT_NAME:
        return jsonify({'message': '默认部门“财务部”不允许删除'}), 409

    in_use = Contract.query.filter(Contract.department == row.name).first()
    if in_use:
        return jsonify({'message': '该部门下已有合同，无法删除'}), 409

    db.session.delete(row)
    db.session.commit()
    return jsonify({'success': True})


@contracts_bp.get('/contracts')
@require_auth
def list_contracts():
    department = (request.args.get('handling_department') or request.args.get('department') or '').strip()
    project = (request.args.get('project') or '').strip()
    status = (request.args.get('status') or '').strip()
    keyword = (request.args.get('keyword') or request.args.get('search') or '').strip()
    has_file = request.args.get('has_file') == 'true'
    is_archived = (request.args.get('is_archived') or '').strip()

    query = Contract.query
    if department == '__empty__':
        query = query.filter(Contract.department.is_(None))
    elif department:
        query = query.filter(Contract.department == department)
    if project == '__empty__':
        query = query.filter(Contract.project.is_(None))
    elif project:
        query = query.filter(Contract.project == project)
    if status:
        query = query.filter(Contract.status == status)
    if has_file:
        query = query.filter(Contract.file_path.isnot(None))
    if is_archived:
        query = query.filter(Contract.is_archived == is_archived)
    if keyword:
        pattern = f'%{keyword}%'
        query = query.filter(or_(
            Contract.contract_number.ilike(pattern),
            Contract.contract_name.ilike(pattern),
            Contract.contract_unit.ilike(pattern),
            Contract.currency.ilike(pattern),
            Contract.handler.ilike(pattern),
            Contract.department.ilike(pattern),
            Contract.contract_determination_method.ilike(pattern),
            Contract.contract_type.ilike(pattern),
            Contract.purchase_type.ilike(pattern),
            Contract.stamp_tax_rate.ilike(pattern),
            Contract.pricing_method.ilike(pattern),
            Contract.save_place.ilike(pattern),
            Contract.is_archived.ilike(pattern),
            Contract.project.ilike(pattern),
            Contract.status.ilike(pattern),
            Contract.file_path.ilike(pattern),
            Contract.fullbody.ilike(pattern),
            Contract.created_by.ilike(pattern),
            cast(Contract.amount, String).ilike(pattern),
            cast(Contract.copy_count, String).ilike(pattern),
            cast(Contract.handling_date, String).ilike(pattern),
            cast(Contract.start_date, String).ilike(pattern),
            cast(Contract.end_date, String).ilike(pattern),
            cast(Contract.created_at, String).ilike(pattern),
            cast(Contract.updated_at, String).ilike(pattern),
        ))

    rows = query.order_by(Contract.updated_at.desc()).all()
    return jsonify([row.to_dict() for row in rows])


@contracts_bp.get('/contracts/statistics')
@require_auth
def get_contract_statistics():
    total_count = Contract.query.count()
    total_amount = db.session.query(func.coalesce(func.sum(Contract.amount), 0)).scalar()

    archived_count = Contract.query.filter(Contract.is_archived == '已归档').count()
    archived_amount = db.session.query(
        func.coalesce(func.sum(Contract.amount), 0)
    ).filter(Contract.is_archived == '已归档').scalar()

    return jsonify({
        'total_count': int(total_count or 0),
        'total_amount': _format_decimal_plain(Decimal(str(total_amount or 0))),
        'archived_count': int(archived_count or 0),
        'archived_amount': _format_decimal_plain(Decimal(str(archived_amount or 0))),
    })


@contracts_bp.get('/contracts/dashboard-charts')
@require_auth
def get_dashboard_charts():
    with_file = Contract.query.filter(Contract.file_path.isnot(None), Contract.file_path != '').count()
    total_contract_count = Contract.query.count()
    contract_file_pie = {
        'with_file': int(with_file),
        'without_file': int(max(total_contract_count - with_file, 0)),
    }

    contract_file_paths = {
        _normalize_contract_file_path(row.file_path)
        for row in Contract.query.filter(Contract.file_path.isnot(None), Contract.file_path != '').all()
        if _normalize_contract_file_path(row.file_path)
    }
    try:
        storage_files = _collect_storage_files()
    except Exception:
        storage_files = []

    storage_paths = {
        _normalize_relative_path(item.get('path') or '')
        for item in storage_files
        if (item.get('path') or '').strip()
    }
    with_contract = len(storage_paths.intersection(contract_file_paths))
    file_contract_pie = {
        'with_contract': int(with_contract),
        'without_contract': int(max(len(storage_paths) - with_contract, 0)),
    }

    departments = _get_department_names()
    contract_counts = []
    file_counts = []
    for name in departments:
        contract_counts.append(Contract.query.filter(Contract.department == name).count())
        file_counts.append(
            Contract.query.filter(
                Contract.department == name,
                Contract.file_path.isnot(None),
                Contract.file_path != '',
            ).count()
        )

    return jsonify({
        'contract_file_pie': contract_file_pie,
        'file_contract_pie': file_contract_pie,
        'dept_bar': {
            'departments': departments,
            'contract_counts': contract_counts,
            'file_counts': file_counts,
        },
    })


@contracts_bp.get('/contracts/<int:contract_id>')
@require_auth
def get_contract(contract_id):
    row = Contract.query.get_or_404(contract_id)
    return jsonify(row.to_dict(include_fullbody=True))


@contracts_bp.post('/contracts/quick-match-files')
@require_auth
def quick_match_contract_files():
    body = request.get_json(silent=True) or {}
    raw_ids = body.get('ids')
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify({'message': 'ids is required'}), 400

    contract_ids = []
    for item in raw_ids:
        try:
            value = int(item)
        except (TypeError, ValueError):
            continue
        if value > 0:
            contract_ids.append(value)

    if not contract_ids:
        return jsonify({'message': 'ids is invalid'}), 400

    try:
        pdf_files = _collect_storage_pdf_files()
    except Exception as exc:
        return jsonify({'message': f'读取存储目录失败: {exc}'}), 500

    results = []
    success_count = 0

    for contract_id in contract_ids:
        row = Contract.query.get(contract_id)
        if not row:
            results.append({
                'id': contract_id,
                'status': 'failed',
                'message': '合同不存在',
            })
            continue

        if (row.is_archived or '').strip() == '已归档':
            results.append({
                'id': row.id,
                'contract_number': row.contract_number,
                'contract_name': row.contract_name,
                'status': 'failed',
                'message': '合同已归档，跳过',
            })
            continue

        if _normalize_contract_file_path(row.file_path):
            results.append({
                'id': row.id,
                'contract_number': row.contract_number,
                'contract_name': row.contract_name,
                'status': 'failed',
                'message': '合同已有附件，跳过',
                'file_path': row.file_path,
            })
            continue

        best_match, matched = _select_best_pdf_match(row.contract_name, pdf_files)
        if not best_match:
            results.append({
                'id': row.id,
                'contract_number': row.contract_number,
                'contract_name': row.contract_name,
                'status': 'failed',
                'message': '未匹配到包含合同名称的PDF文件',
            })
            continue

        row.file_path = best_match['path']
        success_count += 1
        results.append({
            'id': row.id,
            'contract_number': row.contract_number,
            'contract_name': row.contract_name,
            'status': 'success',
            'message': '匹配成功',
            'file_path': best_match['path'],
            'matched_count': len(matched),
            'matched_name': best_match['name'],
            'similarity': round(best_match['similarity'], 6),
        })

    db.session.commit()

    return jsonify({
        'total': len(contract_ids),
        'success': success_count,
        'failed': len(contract_ids) - success_count,
        'results': results,
    })


@contracts_bp.post('/contracts')
@require_auth
def create_contract():
    body = request.get_json(silent=True) or {}

    record, message, status_code, _ = _build_contract_record(body, g.current_user, update_mode=False)
    if record is None:
        return jsonify({'message': message}), status_code

    db.session.add(record)
    db.session.commit()
    return jsonify(record.to_dict(include_fullbody=True)), 201


@contracts_bp.post('/contracts/import-excel')
@require_auth
def import_contracts_excel():
    if 'file' not in request.files:
        return jsonify({'message': 'file is required'}), 400

    uploaded = request.files['file']
    if uploaded.filename == '':
        return jsonify({'message': 'empty filename'}), 400

    file_ext = os.path.splitext(uploaded.filename)[1].lower()
    if file_ext not in EXCEL_ALLOWED_EXTENSIONS:
        return jsonify({'message': '仅支持上传 xls 或 xlsx 文件'}), 400

    try:
        sheet_name, rows = _load_excel_rows(uploaded)
    except ValueError as exc:
        return jsonify({'message': str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({'message': str(exc)}), 500
    except Exception as exc:
        current_app.logger.exception('Excel import failed while reading workbook')
        return jsonify({'message': f'Excel解析失败: {exc}'}), 400

    if not rows:
        return jsonify({'message': 'Excel 文件中没有可读取的数据'}), 400

    header_index, field_indexes, header_labels = _detect_excel_header(rows)
    if not field_indexes:
        return jsonify({'message': '未识别到可用表头，请确认 Excel 中包含合同名称、合同金额、承办部门等列'}), 400

    required_headers = {
        'contract_name': '合同名称',
        'handling_department': '承办部门',
    }
    missing_headers = [label for key, label in required_headers.items() if key not in field_indexes]
    if missing_headers:
        return jsonify({'message': f'Excel 缺少必要列: {", ".join(missing_headers)}'}), 400

    option_sets = _get_contract_option_sets()
    pending_contract_numbers = set()
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    processed_rows = 0
    errors = []
    failed_rows = []
    source_headers = [_stringify_excel_value(cell) for cell in (rows[header_index] if header_index is not None else [])]

    for excel_row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if _is_excel_row_empty(row):
            continue

        payload = _build_import_payload_from_row(row, field_indexes, header_labels, option_sets)
        if not any(str(payload.get(key, '')).strip() for key in CONTRACT_FIELD_KEYS):
            continue

        processed_rows += 1
        record, message, status_code, is_update = _build_contract_record(
            payload,
            g.current_user,
            pending_contract_numbers=pending_contract_numbers,
            update_mode=True,
        )
        if record is None:
            skipped_count += 1
            errors.append({
                'row': excel_row_number,
                'status_code': status_code,
                'message': message,
            })
            failed_rows.append({
                'row': excel_row_number,
                'message': message,
                'row_values': list(row),
            })
            continue

        db.session.add(record)
        if is_update:
            updated_count += 1
        else:
            imported_count += 1
        if record.contract_number:
            pending_contract_numbers.add(record.contract_number)

    if processed_rows == 0:
        return jsonify({'message': 'Excel 表头已识别，但没有可导入的数据行'}), 400

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('Excel import failed while saving rows')
        return jsonify({'message': f'Excel导入保存失败: {exc}'}), 500

    error_report_token, error_report_filename = _store_import_error_report(sheet_name, source_headers, failed_rows)

    return jsonify({
        'sheet_name': sheet_name,
        'header_row': header_index + 1,
        'total_rows': processed_rows,
        'imported_count': imported_count,
        'updated_count': updated_count,
        'skipped_count': skipped_count,
        'errors': errors[:50],
        'error_report_token': error_report_token,
        'error_report_filename': error_report_filename,
    })


@contracts_bp.get('/contracts/import-template')
@require_auth
def download_contract_import_template():
    try:
        output = _build_contract_import_template()
    except Exception as exc:
        current_app.logger.exception('Failed to build contract import template')
        return jsonify({'message': f'导入模板生成失败: {exc}'}), 500

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='合同导入模板.xlsx',
    )


@contracts_bp.get('/contracts/import-error-report/<token>')
@require_auth
def download_contract_import_error_report(token):
    payload = _IMPORT_ERROR_REPORTS.pop(token, None)
    if not payload:
        return jsonify({'message': '导入失败明细不存在或已失效，请重新导入后再下载'}), 404

    return send_file(
        BytesIO(payload['content']),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=payload['filename'],
    )


@contracts_bp.put('/contracts/<int:contract_id>')
@require_auth
def update_contract(contract_id):
    body = request.get_json(silent=True) or {}
    record = Contract.query.get_or_404(contract_id)

    if 'contract_number' in body:
        candidate = (body.get('contract_number') or '').strip()
        if candidate:
            duplicate = Contract.query.filter(
                Contract.contract_number == candidate,
                Contract.id != record.id,
            ).first()
            if duplicate:
                return jsonify({'message': 'contract_number already exists'}), 409
            record.contract_number = candidate
        else:
            record.contract_number = None
    if 'contract_name' in body:
        record.contract_name = (body.get('contract_name') or '').strip() or record.contract_name
    if 'contract_unit' in body:
        record.contract_unit = (body.get('contract_unit') or '').strip() or None



    if 'contract_amount' in body:
        contract_amount_text = str(body.get('contract_amount') or '').strip()
        if contract_amount_text:
            amount = _safe_decimal(contract_amount_text)
            if amount is None:
                return jsonify({'message': 'contract_amount is invalid'}), 400
            record.amount = amount
        else:
            record.amount = None
        current_app.logger.info('调试: contract_amount=%s, amount=%s', body.get('contract_amount'), record.amount)
        
    if 'handler' in body:
        record.handler = (body.get('handler') or '').strip() or None
    if 'handling_department' in body:
        department = (body.get('handling_department') or '').strip()
        if department:
            allowed_departments = _get_department_names()
            if department not in allowed_departments:
                return jsonify({'message': 'handling_department is not in configured department settings'}), 400
            _department_dir(department)
            record.department = department
    if 'contract_determination_method' in body:
        record.contract_determination_method = (body.get('contract_determination_method') or '').strip() or None
    if 'handling_date' in body:
        record.handling_date = _parse_date(body.get('handling_date'))
    if 'contract_type' in body:
        record.contract_type = _normalize_contract_type_value(body.get('contract_type')) or None
        if 'stamp_tax_rate' not in body:
            record.stamp_tax_rate = STAMP_TAX_RATE_BY_CONTRACT_TYPE.get(record.contract_type or '', '') or None
    if 'purchase_type' in body:
        record.purchase_type = (body.get('purchase_type') or '').strip() or None
    if 'stamp_tax_rate' in body:
        record.stamp_tax_rate = (body.get('stamp_tax_rate') or '').strip() or None
    if 'pricing_method' in body:
        record.pricing_method = (body.get('pricing_method') or '').strip() or None
    if 'copy_count' in body:
        copy_count_text = str(body.get('copy_count') or '').strip()
        if copy_count_text:
            if not re.fullmatch(r'\d+', copy_count_text):
                return jsonify({'message': 'copy_count is invalid'}), 400
            record.copy_count = int(copy_count_text)
        else:
            record.copy_count = None
    if 'save_place' in body:
        save_place = (body.get('save_place') or '').strip()
        if len(save_place) > 50:
            return jsonify({'message': 'save_place is too long (max 50)'}), 400
        record.save_place = save_place or None
    if 'is_archived' in body:
        record.is_archived = (body.get('is_archived') or '').strip() or None
    if 'project' in body:
        project = (body.get('project') or '').strip() or None
        if project:
            allowed_projects = _get_project_names()
            if project not in allowed_projects:
                return jsonify({'message': 'project is not in configured project settings'}), 400
        record.project = project
    if 'fullbody' in body:
        record.fullbody = (body.get('fullbody') or '').strip() or None
    if 'file_path' in body:
        normalized_file_path = _normalize_contract_file_path(body.get('file_path'))
        record.file_path = normalized_file_path or None
    if 'start_date' in body:
        record.start_date = _parse_date(body.get('start_date'))
    if 'end_date' in body:
        record.end_date = _parse_date(body.get('end_date'))
    if 'status' in body:
        record.status = (body.get('status') or '').strip() or record.status

    db.session.commit()
    return jsonify(record.to_dict(include_fullbody=True))


@contracts_bp.delete('/contracts/<int:contract_id>')
@require_auth
def delete_contract(contract_id):
    record = Contract.query.get_or_404(contract_id)

    try:
        _delete_contract_file(record)
    except ValueError:
        return jsonify({'message': 'file_path 非法'}), 400
    except Exception as exc:
        return jsonify({'message': f'删除文件失败: {exc}'}), 500

    db.session.delete(record)
    db.session.commit()
    return jsonify({'success': True})


@contracts_bp.post('/contracts/<int:contract_id>/upload')
@require_auth
def upload_contract_file(contract_id):
    record = Contract.query.get_or_404(contract_id)
    if 'file' not in request.files:
        return jsonify({'message': 'file is required'}), 400

    uploaded = request.files['file']
    if uploaded.filename == '':
        return jsonify({'message': 'empty filename'}), 400

    filename = _sanitize_upload_filename(uploaded.filename)

    if current_app.config.get('CONTRACT_STORAGE_MODE') == 'remote':
        remote_folder = _build_filestation_path(record.department)
        try:
            final_name = _synology_upload_file(remote_folder, filename, uploaded)
        except Exception as exc:
            return jsonify({'message': f'远程上传失败: {exc}'}), 500
    else:
        dept_dir = _department_dir(record.department)
        existing_names = [name for name in os.listdir(dept_dir) if os.path.isfile(os.path.join(dept_dir, name))]
        final_name = _next_available_filename(existing_names, filename)
        target_path = os.path.join(dept_dir, final_name)
        uploaded.save(target_path)

    record.file_path = _build_synology_file_path(record.department, final_name)
    db.session.commit()

    return jsonify({'file_path': record.file_path})


def parse_contract_pdf_bak():
    body = request.get_json(silent=True) or {}
    incoming_fullbody = str(body.get('fullbody') or '').strip()
    use_fullbody_directly = len(incoming_fullbody) > 20

    if use_fullbody_directly:
        pdf_text = incoming_fullbody
        preview_lines = _preview_lines(pdf_text)
        current_app.logger.info(
            'AI parse: direct-fullbody mode user=%s chars=%s content_length=%s',
            g.current_user,
            len(pdf_text),
            request.content_length,
        )
    else:
        if 'file' not in request.files:
            return jsonify({'message': 'file is required（或提供长度超过20的fullbody）'}), 400

        uploaded = request.files['file']
        if uploaded.filename == '':
            return jsonify({'message': 'empty filename'}), 400

        current_app.logger.info(
            'AI parse: request user=%s filename=%s mimetype=%s content_length=%s',
            g.current_user,
            uploaded.filename,
            uploaded.mimetype,
            request.content_length,
        )

        filename = (uploaded.filename or '').lower()
        if not filename.endswith('.pdf'):
            return jsonify({'message': '仅支持上传PDF文件'}), 400

        try:
            pdf_text, preview_lines = extract_pdf_text(uploaded)
        except Exception as exc:
            current_app.logger.exception('AI parse: PDF extraction failed')
            return jsonify({'message': f'PDF解析失败: {exc}'}), 400

    if not pdf_text:
        current_app.logger.warning('AI parse: no text extracted, preview_lines=%s', preview_lines)
        return jsonify({
            'message': 'PDF未解析到可用文本，请确认扫描件清晰度/方向或是否含可读文字',
            'ocr_preview_lines': preview_lines,
        }), 400

    try:
        raw_fields = _minimax_extract_fields(pdf_text)
    except Exception as exc:
        current_app.logger.exception('AI parse: Minimax extraction failed')
        return jsonify({'message': f'AI解析失败: {exc}'}), 500

    if not _has_any_field_value(raw_fields):
        current_app.logger.warning('AI parse: extracted fields are all empty')
        return jsonify({
            'message': 'AI返回结果为空，无法自动提取字段，请查看OCR预览并检查PDF清晰度',
            'ocr_preview_lines': preview_lines,
        }), 422

    fields = _normalize_option_fields(raw_fields)
    current_app.logger.info('AI parse: option-normalized fields=%s', fields)

    match_candidates = _find_ai_match_candidates(fields)
    current_app.logger.info('AI parse: matched existing contracts=%s', len(match_candidates))

    current_app.logger.info('AI parse: success extracted fields=%s', fields)

    return jsonify({
        'fields': fields,
        'fullbody': pdf_text,
        'match_candidates': match_candidates,
    })


@contracts_bp.post('/contracts/ai-parse')
@require_auth
def parse_contract_pdf():
    body = request.get_json(silent=True) or {}
    incoming_fullbody = str(body.get('fullbody') or '').strip()
    use_fullbody_directly = len(incoming_fullbody) > 20

    if use_fullbody_directly:
        pdf_text = incoming_fullbody
        preview_lines = _preview_lines(pdf_text)
        current_app.logger.info(
            'AI parse: direct-fullbody mode user=%s chars=%s content_length=%s',
            g.current_user,
            len(pdf_text),
            request.content_length,
        )
    else:
        if 'file' not in request.files:
            return jsonify({'message': 'file is required（或提供长度超过20的fullbody）'}), 400

        uploaded = request.files['file']
        if uploaded.filename == '':
            return jsonify({'message': 'empty filename'}), 400

        current_app.logger.info(
            'AI parse: request user=%s filename=%s mimetype=%s content_length=%s',
            g.current_user,
            uploaded.filename,
            uploaded.mimetype,
            request.content_length,
        )

        filename = (uploaded.filename or '').lower()
        if not filename.endswith('.pdf'):
            return jsonify({'message': '仅支持上传PDF文件'}), 400

        try:
            pdf_text, preview_lines = mineru_extract_text_from_uploaded_pdf(uploaded)
        except Exception as exc:
            current_app.logger.exception('AI parse: PDF extraction failed')
            return jsonify({'message': f'PDF解析失败: {exc}'}), 400

    if not pdf_text:
        current_app.logger.warning('AI parse: no text extracted, preview_lines=%s', preview_lines)
        return jsonify({
            'message': 'PDF未解析到可用文本，请确认扫描件清晰度/方向或是否含可读文字',
            'ocr_preview_lines': preview_lines,
        }), 400

    try:
        raw_fields = _minimax_extract_fields(pdf_text)
    except Exception as exc:
        current_app.logger.exception('AI parse: Minimax extraction failed')
        return jsonify({'message': f'AI解析失败: {exc}'}), 500

    if not _has_any_field_value(raw_fields):
        current_app.logger.warning('AI parse: extracted fields are all empty')
        return jsonify({
            'message': 'AI返回结果为空，无法自动提取字段，请查看OCR预览并检查PDF清晰度',
            'ocr_preview_lines': preview_lines,
        }), 422

    fields = _normalize_option_fields(raw_fields)
    current_app.logger.info('AI parse: option-normalized fields=%s', fields)

    match_candidates = _find_ai_match_candidates(fields)
    current_app.logger.info('AI parse: matched existing contracts=%s', len(match_candidates))

    current_app.logger.info('AI parse: success extracted fields=%s', fields)

    return jsonify({
        'fields': fields,
        'fullbody': pdf_text,
        'match_candidates': match_candidates,
    })


@contracts_bp.get('/contracts/<int:contract_id>/download')
@require_auth
def download_contract_file(contract_id):
    record = Contract.query.get_or_404(contract_id)
    try:
        content, file_name, mime = _load_contract_file_payload(record)
    except PermissionError as exc:
        return jsonify({'message': str(exc)}), 401
    except ValueError:
        return jsonify({'message': 'file_path 非法'}), 400
    except FileNotFoundError as exc:
        return jsonify({'message': str(exc)}), 404
    except Exception as exc:
        return jsonify({'message': f'文件下载失败: {exc}'}), 500

    return send_file(
        BytesIO(content),
        mimetype=mime,
        as_attachment=True,
        download_name=file_name,
    )


@contracts_bp.get('/contracts/<int:contract_id>/preview')
@require_auth
def preview_contract_file(contract_id):
    record = Contract.query.get_or_404(contract_id)
    normalized_file_path = _normalize_contract_file_path(record.file_path)
    if not normalized_file_path:
        return jsonify({'message': '该合同未上传文件'}), 404

    file_name = os.path.basename(normalized_file_path)
    if not file_name.lower().endswith('.pdf'):
        return jsonify({'message': '仅支持PDF预览'}), 400

    try:
        content, payload_name, _mime = _load_contract_file_payload(record)
    except PermissionError as exc:
        return jsonify({'message': str(exc)}), 401
    except ValueError:
        return jsonify({'message': 'file_path 非法'}), 400
    except FileNotFoundError as exc:
        return jsonify({'message': str(exc)}), 404
    except Exception as exc:
        return jsonify({'message': f'文件预览失败: {exc}'}), 500

    final_name = payload_name or file_name or f'contract_{record.id}.pdf'
    return send_file(
        BytesIO(content),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=final_name,
    )
